import json, csv, os, collections
DATE='2026-09-05'
b=json.load(open('/tmp/built.json'))
overrides=b['overrides']; rows=b['rows']; counts=b['counts']; log=b['search_log']
cohort=json.load(open('data/research/residents/pass_14_76_cohort.json'))
order=[p['person_id'] for p in cohort['people']]

f=json.load(open('data/research/residents/pass_14_findings.json'))
f['_doc']=("Completed T-0509 ledger — cohort 14 of the resident-research programme. Every one of the 76 frozen-manifest "
  "members received a dated review on 2026-09-05 against the repository corpus and its committed crosswalks, and the "
  "outcome is written here with the discriminator, or its absence, that decided it. Later-volume readings are recorded "
  "in `notes`, which the synthesizer does not promote canonical facts out of: back-projecting a trade or an age across "
  "eight years is T-0514/T-0515's decision, not this pass's.")
f['reviewed_on']=DATE
f['status']='complete'
f['method']=("Exact name first, then justified initial and OCR variants, across the committed newspapers and identity "
  "ledger, the 1833 tax and 1834/1835 poll lists, the Illinois public-domain land tract sales, the 1830 census of the "
  "Chicago precinct, the 1840 census heads, the St Cyr and St Mary registers, Fergus 1839, Fergus 1843, Norris 1844, "
  "Fergus's Historical Series 26-29 (the 1843 directory, its advertising pages, its fire-department and civic rolls, "
  "and Wentworth's obituary lists), the Calumet Club old-settler rolls, the Newberry genealogical index cards and the "
  "Genealogy Trails transcriptions. Where the repository already holds a crosswalk verdict for a name, that verdict is "
  "quoted rather than re-decided; what this pass adds on top is a directed reading of Fergus 26-29 against all 76 names, "
  "which the directory claim ledgers do not cover.")
f['outcome_rule']=("corroborated = an agreement, forename for forename or initial for initial, with an independent source "
  "written at or before the scene year; corroborated_enrichment = the same agreement but only in a volume printed after "
  "1835, which enriches biography and adds no 1835 attestation; candidate_identity = a plausible external identity with "
  "no date, place, occupation or kinship discriminator bridging it; no_corroboration = the post-office lists and a "
  "documented refusal, which is a negative search and not evidence of absence.")
f['outcome_counts']=dict(sorted(counts.items()))
f['completed_person_ids']=order
f['pending_person_ids']=[]
f.pop('pending', None)
f['overrides']={pid: overrides[pid] for pid in order}
json.dump(f, open('data/research/residents/pass_14_findings.json','w'), indent=1, ensure_ascii=False)
f2=open('data/research/residents/pass_14_findings.json','a'); f2.write('\n'); f2.close()

PKG='../reference/resident-research/T-0509'
os.makedirs(PKG, exist_ok=True)
HEAD=['person_id','name_transcribed','name_normalized','stratum','outcome','candidate_ids','proposed_facts',
      'evidence_for','evidence_against','source_ids','source_urls_tiers','queries','access_date','notes']
byid={r['person_id']: r for r in rows}

# the hand candidates this pass minted, with the discriminator that refused each one
CANDIDATES=[
 ('cand_curtiss_j_fergus_1843_james_attorney','curtiss_j',
  "Fergus 1843: 'Curtiss, James, State's attorney, 136 Lake, res W. Randolph, bet May and Ann [9th mayor, died, Joliet, Ill., November 2, 1859, aged 56'.",
  "The 1835 record carries the initial J. alone, and the same volume prints a second J. Curtiss."),
 ('cand_curtiss_j_fergus_1843_jw_gunsmith','curtiss_j',
  "Fergus 1843: 'Curtiss. J. W., gunsmith, res cor North Water and Wolcott'.",
  "The 1835 record carries the initial J. alone, and the same volume prints James Curtiss the attorney as its rival."),
 ('cand_crocker_h_fergus_hans_lawyer','crocker_h',
  "Fergus 26-29 obituary list: 'Crocker. Hans, lawyer, died, Milwaukee, Wis., Mar. 17, 1889, a. 73'.",
  "The age puts birth about 1816, which would make him nineteen at the scene date, and the death is at Milwaukee."),
]
with open(f'{PKG}/T-0509_resident_research.csv','w',newline='',encoding='utf-8') as fh:
    w=csv.DictWriter(fh, fieldnames=HEAD); w.writeheader()
    for pid in order: w.writerow(byid[pid])
print('csv rows', len(order))

try:
    import openpyxl
    from openpyxl import Workbook
    wb=Workbook(); ws=wb.active; ws.title='Residents'
    ws.append(HEAD)
    for pid in order: ws.append([byid[pid][h] for h in HEAD])
    wc=wb.create_sheet('Candidates')
    wc.append(['candidate_id','person_id','name','asserted','basis','conflicts','source_ids'])
    for cid, pid, basis, conflict in CANDIDATES:
        wc.append([cid, pid, byid[pid]['name_transcribed'], 'FALSE', basis, conflict, 'fergus_historical_series_26_29'])
    for pid in [r['person_id'] for r in rows if r['outcome']=='candidate_identity' and not r['candidate_ids']]:
        wc.append([f'cand_{pid}_repository_lead', pid, byid[pid]['name_transcribed'], 'FALSE',
          "A repository crosswalk holds a contested lead on this name: an external record the corpus could not refuse outright.",
          "No date, place, occupation or kinship discriminator bridges the lead to the 1835 person, so it is retained as answerable and not as an identity.",
          ''])
    wsrc=wb.create_sheet('Sources')
    wsrc.append(['source_id','times_cited','resolves_in_data_sources'])
    cnt=collections.Counter()
    for r in rows:
        for s in (r['source_ids'] or '').split(';'):
            if s: cnt[s]+=1
    for s,n in cnt.most_common():
        wsrc.append([s,n,os.path.exists(f'data/sources/{s}.json')])
    wl=wb.create_sheet('Search_Log')
    wl.append(['person_id','name','date','corpora_queried','result','limitations'])
    for e in log:
        wl.append([e['person_id'],e['name'],e['date'],e['corpora'],e['result'],e['limitation']])
    wb.save(f'{PKG}/T-0509_resident_research_working.xlsx')
    print('xlsx written; search_log rows', len(log))
except ImportError:
    print('openpyxl missing — CSV only')
