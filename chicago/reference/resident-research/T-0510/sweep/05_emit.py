import json, csv, os, collections
DATE='2026-09-05'
b=json.load(open('/tmp/built.json'))
overrides=b['overrides']; rows=b['rows']; counts=b['counts']; log=b['search_log']
cohort=json.load(open('data/research/residents/pass_15_76_cohort.json'))
order=[p['person_id'] for p in cohort['people']]

f=json.load(open('data/research/residents/pass_15_findings.json'))
f['_doc']=("Completed T-0510 ledger — cohort 15 of the resident-research programme. Every one of the 76 frozen-manifest "
  "members received a dated review on 2026-09-05 against the repository corpus and its committed crosswalks, and the "
  "outcome is written here with the discriminator, or its absence, that decided it. Later-volume readings are recorded "
  "in `notes`, which the synthesizer does not promote canonical facts out of: back-projecting a trade or an age across "
  "eight years is T-0514/T-0515's decision, not this pass's.")
f['reviewed_on']=DATE
f['status']='complete'
f['method']=("Exact name first, then justified initial and OCR variants, across the committed newspapers and identity "
  "ledger, the 1833 tax and 1834/1835 poll lists, the Illinois public-domain land tract sales, the 1830 census of the "
  "Chicago precinct, the 1840 census heads, the St Cyr and St Mary registers, Fergus 1839, Fergus 1843, Norris 1844, "
  "Fergus's Historical Series 26-29 (the 1843 directory and Wentworth's obituary lists), the Calumet Club old-settler "
  "rolls, the Newberry genealogical index cards and the Genealogy Trails transcriptions. Where the repository already "
  "holds a crosswalk verdict for a name, that verdict is quoted rather than re-decided.")
f['outcome_rule']=("corroborated = an agreement, forename for forename or initial for initial, with an independent source "
  "written at or before the scene year; corroborated_enrichment = the same agreement but only in a volume printed after "
  "1835, which enriches biography and adds no 1835 attestation; candidate_identity = a plausible external identity with "
  "no date, place, occupation or kinship discriminator bridging it; no_corroboration = the post-office lists and a "
  "documented refusal, which is a negative search and not evidence of absence.")
f['outcome_counts']=dict(sorted(counts.items()))
f['completed_person_ids']=order
f['pending_person_ids']=[]
f.pop('pending', None)
f['overrides']={pid: overrides[pid] for pid in order}
json.dump(f, open('data/research/residents/pass_15_findings.json','w'), indent=1, ensure_ascii=False)
f2=open('data/research/residents/pass_15_findings.json','a'); f2.write('\n'); f2.close()

PKG='../reference/resident-research/T-0510'
os.makedirs(PKG, exist_ok=True)
HEAD=['person_id','name_transcribed','name_normalized','stratum','outcome','candidate_ids','proposed_facts',
      'evidence_for','evidence_against','source_ids','source_urls_tiers','queries','access_date','notes']
byid={r['person_id']: r for r in rows}
with open(f'{PKG}/T-0510_resident_research.csv','w',newline='',encoding='utf-8') as fh:
    w=csv.DictWriter(fh, fieldnames=HEAD); w.writeheader()
    for pid in order: w.writerow(byid[pid])
print('csv rows', len(order))

try:
    import openpyxl
    from openpyxl import Workbook
    wb=Workbook(); ws=wb.active; ws.title='Residents'
    ws.append(HEAD)
    for pid in order: ws.append([byid[pid][h] for h in HEAD])
    wc=wb.create_sheet('Candidates')
    wc.append(['candidate_id','person_id','name','asserted','basis','conflicts','source_ids'])
    wc.append(['cand_thomas_hoit_fergus_1843_carpenter','hoit_thomas','Thomas Hoit, carpenter (Fergus 1843)','FALSE',
      "Exact full name in the 1843 Chicago Directory, same town, eight years after the scene date.",
      "The bracketed death note reads 'died 1881, aged 60', which puts birth about 1821 and would make him fourteen when the Democrat printed the 1835 resident.",
      'fergus_historical_series_26_29'])
    for pid,label in (('dow_albert_f','Albert F. Dow'),('ford_ebenezer','Ebenezer Ford'),('porter_eliza_chappel','Eliza Chappel Porter')):
        wc.append([f'cand_{pid}_newberry_surname_card', pid, label, 'FALSE',
          'Newberry genealogical index cards heading this surname name Chicago or Cook County and cite works the project can test.',
          "A surname-only index heading is a refusal under this project's standing rule; the lead is retained as answerable, not as an identity.",
          'newberry_genealogical_index'])
    wsrc=wb.create_sheet('Sources')
    wsrc.append(['source_id','times_cited','resolves_in_data_sources'])
    cnt=collections.Counter()
    for r in rows:
        for s in (r['source_ids'] or '').split(';'):
            if s: cnt[s]+=1
    for s,n in cnt.most_common():
        wsrc.append([s,n,os.path.exists(f'data/sources/{s}.json')])
    wl=wb.create_sheet('Search_Log')
    wl.append(['person_id','name','date','corpora_queried','result','limitations'])
    for e in log:
        wl.append([e['person_id'],e['name'],e['date'],e['corpora'],e['result'],e['limitation']])
    wb.save(f'{PKG}/T-0510_resident_research_working.xlsx')
    print('xlsx written; search_log rows', len(log))
except ImportError:
    print('openpyxl missing — CSV only')
