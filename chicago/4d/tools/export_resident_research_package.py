#!/usr/bin/env python3
"""Write a resident-research cohort's durable reference package, or check one.

T-0511.  `chicago/reference/resident-research/README.md` says a cohort ticket is
not complete "while its XLSX/CSV/README package exists only locally", and on dev
the folders existed for T-0478..T-0486 only.  The pilot (T-0442) and passes 2 and
3 (T-0462, T-0463) — the first 225 people, three of the "11 slices" the owner is
counting — had findings JSON and dossiers and nothing a reader can open.

The later cohorts each wrote their package from a one-off script
(`complete_resident_research_pass_5.py`, `pack_t0493_reference.py`).  This is the
same job written once, from the two committed records that already hold every
field the package needs:

  * `data/research/residents/<cohort>.json` — the FROZEN manifest: who was in the
    cohort, the evidence each person started from, the seed sources.
  * `data/residents/research_pilot.json` — the committed public review payload:
    outcome, summary, queries, sources and candidates, one review per person, in
    pass order.  `compile_resident_research_pilot.py --gate` re-derives it, so
    what this tool exports is gated upstream of itself.

Nothing here researches anything, alters an outcome or mints a person: it is a
projection of records that already exist into the shape the README contracts for.

    python3 tools/export_resident_research_package.py T-0442 --build
    python3 tools/export_resident_research_package.py --all --build
    python3 tools/export_resident_research_package.py --check     # gate
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
CHICAGO = ROOT.parent
RESEARCH = ROOT / "data" / "research" / "residents"
RESIDENTS = ROOT / "data" / "residents"
SOURCES = ROOT / "data" / "sources"
REFERENCE = CHICAGO / "reference" / "resident-research"
PAYLOAD = RESIDENTS / "research_pilot.json"

# The three cohorts this tool exports, and where each one's people sit in the
# review payload.  `offset` is not a guess: the payload lists its passes in order
# in `passes[]`, and --build asserts the slice it takes against that list before
# it writes a byte.
COHORTS = {
    "T-0442": {
        "cohort": "1",
        "label": "pilot",
        "manifest": "pilot_75_cohort.json",
        "findings": None,
        "selector": "tools/select_resident_research_pilot.py",
        "dossier": "docs/RESEARCH/resident_identity_pilot_75.md",
    },
    "T-0462": {
        "cohort": "2",
        "label": "pass 2",
        "manifest": "pass_02_75_cohort.json",
        "findings": "pass_02_findings.json",
        "selector": "tools/select_resident_research_pass_2.py",
        "dossier": "docs/RESEARCH/resident_identity_pass_02_75.md",
    },
    "T-0463": {
        "cohort": "3",
        "label": "pass 3",
        "manifest": "pass_03_75_cohort.json",
        "findings": "pass_03_findings.json",
        "selector": "tools/select_resident_research_pass_3.py",
        "dossier": "docs/RESEARCH/resident_identity_pass_03_75.md",
    },
}

# Every cohort ticket that has completed research and therefore owes a package.
# --check refuses a folder that has gone missing as loudly as a stale one; the
# completion rule in the reference README is what is being gated.
PACKAGES_OWED = [
    "T-0442", "T-0462", "T-0463", "T-0478", "T-0479", "T-0480", "T-0481",
    "T-0482", "T-0483", "T-0484", "T-0485", "T-0486", "T-0493",
]

HEADERS = [
    "ticket", "cohort", "person_id", "household_id", "name_transcribed", "name_normalized",
    "stratum", "seed_source_id", "seed_source_date", "letter_list_dates", "research_outcome",
    "identity_confidence", "residence_confidence", "household_confidence", "occupation_confidence",
    "candidate_ids", "proposed_birth", "proposed_death", "proposed_arrival_migration",
    "proposed_occupation_trade", "proposed_address_property", "proposed_spouse_kin",
    "proposed_civic_voter_census", "proposed_household_facts", "evidence_for", "evidence_against",
    "source_ids", "source_urls_locators", "source_tiers", "queries", "access_dates",
    "source_limitations", "recommended_data_action", "notes",
]

IDENTITY_CONFIDENCE = {
    "corroborated_enrichment": "corroborated",
    "candidate_identity": "candidate — unasserted",
    "no_corroboration": "unresolved / no safe external match",
}
ACTION = {
    "corroborated_enrichment": "T-0487 adjudication; T-0488 attested promotion where independently supported",
    "candidate_identity": "retain candidate unasserted; resolve in T-0487 before any canonical identity change",
    "no_corroboration": "retain documented no-find; revisit only with new evidence",
}
LIMITATION = (
    "Candidate and no-find decisions are bounded by the sources searched; name "
    "similarity alone was not accepted as identity evidence."
)


def load(path: Path):
    return json.loads(path.read_text())


def slug(text: str) -> str:
    return re.sub(r"_+", "_", re.sub(r"[^a-z0-9]+", "_", (text or "").lower())).strip("_")


def candidate_id(person_id: str, candidate: dict) -> str:
    """A stable row key for a candidate the payload gives no id.

    Name-scoped like T-0479's hand-written ids, person-scoped so two people may
    consider the same outside name without colliding.  It keys a research row and
    asserts nothing: `asserted` stays whatever the payload says it is.
    """
    return "cand_%s__%s" % (slug(candidate.get("name", "")) or "unnamed", person_id)


def person_index() -> dict:
    idx = load(RESIDENTS / "index.json")
    people = {}
    for entry in idx["households"]:
        household = load(RESIDENTS / entry["file"])
        for person in household.get("persons", []):
            people[person["id"]] = (household, person)
    return people


def dated(value) -> str:
    """`{value, confidence}` sidecars read as text; a bare scalar reads as itself."""
    if isinstance(value, dict):
        return ("%s (%s)" % (value.get("value", ""), value.get("confidence", ""))).strip()
    return "" if value is None else str(value)


def joined(values) -> str:
    if not values:
        return ""
    if not isinstance(values, list):
        return str(values)
    out = []
    for value in values:
        out.append(json.dumps(value, ensure_ascii=False, sort_keys=True)
                   if isinstance(value, (dict, list)) else str(value))
    return "; ".join(out)


def reviews_for(ticket: str, payload: dict) -> list[dict]:
    """The payload's slice for one ticket, taken from its own `passes[]` order."""
    offset = 0
    for entry in payload["passes"]:
        if entry["ticket"] == ticket:
            reviews = payload["reviews"][offset:offset + entry["size"]]
            counts = dict(Counter(r["outcome"] for r in reviews))
            if len(reviews) != entry["size"] or counts != entry["counts"]:
                raise SystemExit(
                    "%s: the payload slice does not match its own pass record "
                    "(rows=%d counts=%s)" % (ticket, len(reviews), counts))
            return reviews
        offset += entry["size"]
    raise SystemExit("%s: no pass by that name in %s" % (ticket, PAYLOAD.name))


def build_rows(ticket: str) -> tuple[list[dict], list[list], list[list], list[list], dict]:
    spec = COHORTS[ticket]
    manifest = load(RESEARCH / spec["manifest"])
    payload = load(PAYLOAD)
    reviews = {r["person_id"]: r for r in reviews_for(ticket, payload)}
    members = manifest["people"]
    if {m["person_id"] for m in members} != set(reviews):
        raise SystemExit("%s: the frozen manifest and the review payload name different people" % ticket)

    people = person_index()
    source_cache = {p.stem: load(p) for p in SOURCES.glob("*.json")}

    rows, candidate_rows, search_rows = [], [], []
    used_sources: dict[str, dict] = {}

    for member in members:
        pid = member["person_id"]
        review = reviews[pid]
        if pid not in people:
            raise SystemExit("%s: %s is in the cohort and not in the residents layer" % (ticket, pid))
        household, person = people[pid]
        outcome = review["outcome"]
        if outcome not in IDENTITY_CONFIDENCE:
            raise SystemExit("%s: %s carries an unsupported outcome %r" % (ticket, pid, outcome))

        source_ids = review.get("sources") or []
        locators, tiers, limitations = [], [], []
        for sid in source_ids:
            source = source_cache.get(sid)
            if not source:
                raise SystemExit("%s: %s cites a source with no record: %s" % (ticket, pid, sid))
            used_sources[sid] = source
            locator = source.get("url", "")
            if source.get("locator"):
                locator += " — " + source["locator"]
            locators.append("%s: %s" % (sid, locator))
            tiers.append("%s:%s" % (sid, source.get("tier", "")))
            if source.get("note"):
                limitations.append("%s: %s" % (sid, source["note"]))

        candidates = review.get("candidates") or []
        ids_here = []
        for candidate in candidates:
            cid = candidate_id(pid, candidate)
            ids_here.append(cid)
            for sid in candidate.get("sources") or []:
                if sid in source_cache:
                    used_sources[sid] = source_cache[sid]
            candidate_rows.append([
                cid, pid, person["name"], candidate.get("name", ""),
                candidate.get("assessment", ""),
                "yes" if candidate.get("asserted") else "no",
                candidate.get("basis", ""),
                " | ".join(candidate.get("conflicts") or []),
                "; ".join(candidate.get("sources") or []),
            ])

        queries = review.get("queries") or []
        for query in queries:
            search_rows.append([
                pid, person["name"], query, review.get("reviewed_on", ""), outcome,
                "; ".join(source_ids), LIMITATION,
            ])

        row = {h: "" for h in HEADERS}
        row.update({
            "ticket": ticket,
            "cohort": spec["cohort"],
            "person_id": pid,
            "household_id": household["id"],
            "name_transcribed": member.get("name") or person["name"],
            "name_normalized": person["name"],
            "stratum": member.get("starting_evidence", ""),
            "seed_source_id": joined(member.get("sources") or person.get("sources") or []),
            "letter_list_dates": joined(member.get("letter_list_returns") or []),
            "research_outcome": outcome,
            "identity_confidence": IDENTITY_CONFIDENCE[outcome],
            "residence_confidence": dated(household.get("present_on_scene_date")),
            "household_confidence": (
                "existing canonical household; composition not re-adjudicated in %s" % ticket),
            "occupation_confidence": dated(person.get("occupation")),
            "candidate_ids": "; ".join(ids_here),
            "evidence_for": review.get("summary", ""),
            "evidence_against": " | ".join(
                "; ".join(c.get("conflicts") or []) for c in candidates if c.get("conflicts")),
            "source_ids": "; ".join(source_ids),
            "source_urls_locators": " | ".join(locators),
            "source_tiers": "; ".join(tiers),
            "queries": " | ".join(queries),
            "access_dates": review.get("reviewed_on", ""),
            "source_limitations": " | ".join(limitations) or LIMITATION,
            "recommended_data_action": ACTION[outcome],
            "notes": review.get("identity_rule", ""),
        })
        rows.append(row)

    source_rows = []
    for sid in sorted(used_sources):
        doc = used_sources[sid]
        source_rows.append([
            sid, doc.get("tier", ""), doc.get("title", "") or doc.get("citation", ""),
            doc.get("url", ""), doc.get("locator", ""), doc.get("note", ""),
        ])

    counts = dict(Counter(r["research_outcome"] for r in rows))
    summary = {
        "ticket": ticket,
        "label": spec["label"],
        "size": len(rows),
        "counts": counts,
        "reviewed_on": sorted({r["access_dates"] for r in rows if r["access_dates"]}),
        "strata": dict(Counter(r["stratum"] for r in rows)),
        "candidates": len(candidate_rows),
        "sources": len(source_rows),
        "searches": len(search_rows),
        "frame": manifest.get("population_frame", {}),
    }
    return rows, candidate_rows, source_rows, search_rows, summary


CANDIDATE_HEADER = ["candidate_id", "person_id", "resident_name", "candidate_name",
                    "assessment", "asserted", "basis", "conflicts", "source_ids"]
SOURCE_HEADER = ["source_id", "tier", "title_or_citation", "url", "locator", "note"]
SEARCH_HEADER = ["person_id", "resident_name", "query", "searched_on", "result",
                 "source_ids", "limitations"]


def readme_text(ticket: str, summary: dict) -> str:
    spec = COHORTS[ticket]
    counts = summary["counts"]
    frame = summary["frame"]
    reviewed = ", ".join(summary["reviewed_on"]) or "not recorded"
    # The selector's own frame names the strata it sampled — the pilot drew its
    # letter-list half in two, present and uncertain, and the row's
    # `starting_evidence` collapses them back into one. Prefer the frame, and fall
    # back to the counted column only if it has stopped adding up to the cohort.
    breakdown = frame.get("strata") or {}
    if sum(breakdown.values()) != summary["size"]:
        breakdown = summary["strata"]
    strata = "\n".join(
        "- %d %s" % (n, name.replace("_", " "))
        for name, n in sorted(breakdown.items(), key=lambda kv: (-kv[1], kv[0])))
    prior = frame.get("previously_reviewed", 0)
    return """# %(ticket)s — resident research %(label)s

## Scope

This folder is the durable, human-reviewable handoff for **%(ticket)s**, the %(label)s
cohort of the Chicago resident-identity programme, for the 1835-07-01 scene date. It was
written by `tools/export_resident_research_package.py` from the committed records the pass
already left, and it asserts nothing they do not: no outcome is altered here and no person
is minted here.

The cohort is fixed at %(size)d unique, named, non-reconstructed people:

%(strata)s

Of an eligible population of %(eligible)s named non-reconstructed people, %(prior)d had been
reviewed before this pass and %(cumulative)d after it.

## Results

- **%(corroborated)d corroborated enrichments**
- **%(candidate)d candidate identities / duplicate leads**, explicitly unasserted
- **%(nofind)d documented no-corroboration outcomes**
- **0 pending**

Reviewed %(reviewed)s. A documented no-find is a research result, not evidence that the
person did not exist. An exact name is a search lead, not an identity assertion: a candidate
is bridged to the 1835 Chicago record by more than name similarity, or it stays a candidate.

## Confidence rules

Every row carries the outcome the review payload records and the confidence the canonical
record carries, side by side, and never a promotion of one into the other:

- `research_outcome` — `corroborated_enrichment`, `candidate_identity` or `no_corroboration`,
  as the pass resolved it.
- `identity_confidence` — the same judgement in words; `candidate — unasserted` never means
  the identity was accepted.
- `residence_confidence`, `occupation_confidence` — the household's and person's own
  `{value, confidence}` sidecars as they stand in the residents layer, not this pass's opinion.
- `household_confidence` — composition was not re-adjudicated by a research pass.

Competing geography, later-dated records and rejected leads are kept in the Candidates sheet
with their conflicts rather than normalized away.

## Files

- `%(ticket)s_resident_research.csv` — machine-readable Residents table on the shared
  template header (`../cohort_research_template.csv`).
- `%(ticket)s_resident_research_working.xlsx` — working workbook: Residents, Candidates,
  Sources, Search_Log and Summary sheets. Written when `openpyxl` imports.
- `README.md` — this note.

Counts in this package: %(size)d residents, %(ncand)d candidate rows, %(nsrc)d sources,
%(nsearch)d logged searches.

## The records this is derived from

- `chicago/4d/data/research/residents/%(manifest)s` — the frozen cohort manifest, re-derived
  by `%(selector)s --gate`.
- `chicago/4d/data/residents/research_pilot.json` — the public review payload holding this
  pass's outcomes, summaries, queries, sources and candidates; re-derived by
  `tools/compile_resident_research_pilot.py --gate`.%(findings)s
- `chicago/4d/%(dossier)s` — the pass's narrative dossier.
- `chicago/4d/data/sources/*.json` — the stable source records the rows cite.

`python3 tools/export_resident_research_package.py %(ticket)s --check` re-derives this folder
and fails if it has drifted from those records; `tools/check.sh` runs it.

## Unresolved

The %(candidate)d candidate identities are unresolved by design and are handed to the
T-0487–T-0490 adjudication sequence, not promoted here. The %(nofind)d no-corroboration
outcomes are receipts of a bounded search, revisited only when new evidence arrives.
""" % {
        "ticket": ticket,
        "label": spec["label"],
        "size": summary["size"],
        "strata": strata,
        "eligible": frame.get("eligible_real_named_people", "an unrecorded number of"),
        "prior": prior,
        "cumulative": frame.get("cumulative_reviewed", prior + summary["size"]),
        "corroborated": counts.get("corroborated_enrichment", 0),
        "candidate": counts.get("candidate_identity", 0),
        "nofind": counts.get("no_corroboration", 0),
        "reviewed": reviewed,
        "ncand": summary["candidates"],
        "nsrc": summary["sources"],
        "nsearch": summary["searches"],
        "manifest": spec["manifest"],
        "selector": spec["selector"],
        "findings": ("\n- `chicago/4d/data/research/residents/%s` — the pass's authoritative "
                     "outcome/candidate ledger." % spec["findings"]) if spec["findings"] else "",
        "dossier": spec["dossier"],
    }


# ---------------------------------------------------------------- the index
#
# T-0511 asked for one place a reader can see the eleven slices. It is DERIVED
# from the packages on disk rather than typed, because a hand-kept index is the
# first thing to stop being true. The older packages were each written by their
# own one-off script and do not agree on a column name or an outcome spelling,
# so the reader below is deliberately tolerant: it normalizes what it finds and
# says `—` where a package recorded nothing, rather than inventing a number.

OUTCOME_COLUMNS = ("research_outcome", "outcome", "result")
DATE_COLUMNS = ("access_dates", "access_date", "searched_on", "reviewed_on")
OUTCOME_ALIASES = {
    "corroborated": "corroborated_enrichment",
    "corroborated_enrichment": "corroborated_enrichment",
    "candidate": "candidate_identity",
    "candidate_identity": "candidate_identity",
    "no_corroboration": "no_corroboration",
    "no_corroboration_yet": "no_corroboration",
}
INDEX_START = "<!-- index:start — generated by tools/export_resident_research_package.py -->"
INDEX_END = "<!-- index:end -->"


def read_package(ticket: str) -> dict | None:
    folder = REFERENCE / ticket
    files = sorted(folder.glob("*_resident_research.csv"))
    if not files:
        return None
    with files[0].open(newline="", encoding="utf-8-sig") as fh:
        rows = list(csv.DictReader(fh))
    counts: Counter = Counter()
    dates: set[str] = set()
    for row in rows:
        for column in OUTCOME_COLUMNS:
            value = (row.get(column) or "").strip()
            if value:
                counts[OUTCOME_ALIASES.get(value, value)] += 1
                break
        for column in DATE_COLUMNS:
            value = (row.get(column) or "").strip()
            if value:
                dates.add(value.split()[0])
                break
    return {
        "ticket": ticket,
        "rows": len(rows),
        "counts": counts,
        "dates": sorted(dates),
        "xlsx": bool(list(folder.glob("*.xlsx"))),
        "readme": (folder / "README.md").exists(),
    }


def index_rows() -> list[str]:
    lines = [
        "| package | cohort | rows | corroborated | candidate | no-corroboration | recorded | researched |",
        "| --- | --- | ---: | ---: | ---: | ---: | --- | --- |",
    ]
    total = Counter()
    for ticket in indexed_packages():
        pkg = read_package(ticket)
        if pkg is None:
            lines.append("| `%s/` | — | — | — | — | — | **missing** | — |" % ticket)
            continue
        counts = pkg["counts"]
        total["rows"] += pkg["rows"]
        for key in ("corroborated_enrichment", "candidate_identity", "no_corroboration"):
            total[key] += counts.get(key, 0)
        unrecorded = pkg["rows"] - sum(counts.get(k, 0) for k in
                                       ("corroborated_enrichment", "candidate_identity", "no_corroboration"))
        cell = lambda key: str(counts.get(key, 0)) if not unrecorded else (
            str(counts.get(key, 0)) if counts.get(key) else "—")
        dates = ", ".join(pkg["dates"]) or "—"
        held = "CSV" + (" · XLSX" if pkg["xlsx"] else "") + (" · README" if pkg["readme"] else "")
        label = COHORTS[ticket]["label"] if ticket in COHORTS else PACKAGE_LABELS.get(ticket, "")
        lines.append("| `%s/` | %s | %d | %s | %s | %s | %s | %s |" % (
            ticket, label, pkg["rows"], cell("corroborated_enrichment"),
            cell("candidate_identity"), cell("no_corroboration"), held, dates))
    lines.append("| **total** | | **%d** | **%d** | **%d** | **%d** | | |" % (
        total["rows"], total["corroborated_enrichment"], total["candidate_identity"],
        total["no_corroboration"]))
    audit = audit_row()
    if audit is not None:
        lines.append(audit)
    return lines


PACKAGE_LABELS = {
    "T-0478": "pass 4", "T-0479": "pass 5", "T-0480": "pass 6", "T-0481": "pass 7",
    "T-0482": "pass 8", "T-0483": "pass 9", "T-0484": "pass 10", "T-0485": "pass 11",
    "T-0486": "pass 12", "T-0493": "the 1833-35 voter lists", "T-0508": "cohort 13",
    "T-0509": "cohort 14", "T-0510": "cohort 15",
}

PACKAGE_DIR_RE = re.compile(r"^T-\d{4}$")


def indexed_packages() -> list[str]:
    """Every package the index names: the ones the completion rule OWES, plus every
    package folder actually on disk.

    T-0511 hard-coded the owed list, which was right for the eleven slices it was
    written for and wrong the moment cohort 15 landed (T-0518): T-0510 had a full
    package on disk and the index did not know it existed. The owed list still says
    what MUST be there — that is the completion assertion — but what the index SHOWS
    is read off the directory, so a new cohort's package cannot be invisible.

    The cost of that is a coupling worth stating: a new package folder makes
    `--check` report drift until `--index` is re-run. That is the intended failure —
    it is one command, and it is louder than a package nobody can find.
    """
    found = set(PACKAGES_OWED)
    if REFERENCE.is_dir():
        for folder in REFERENCE.iterdir():
            if folder.is_dir() and PACKAGE_DIR_RE.match(folder.name) \
                    and list(folder.glob("*_resident_research.csv")):
                found.add(folder.name)
    return sorted(found)


def audit_row() -> str | None:
    """The programme audit is not a cohort package — it is the whole layer, one row
    per person, generated by `export_resident_audit.py`. The index carries it because
    a reader looking for "what did the programme reach" should not have to know that
    it lives in a differently-shaped folder."""
    folder = REFERENCE / "final" / "audit"
    csvs = sorted(folder.glob("*.csv"))
    if not csvs:
        return None
    with csvs[0].open(newline="", encoding="utf-8-sig") as fh:
        rows = list(csv.DictReader(fh))
    reviewed = sum(1 for r in rows if (r.get("research_ticket") or "").strip())
    held = "CSV" + (" · XLSX" if list(folder.glob("*.xlsx")) else "") \
        + (" · README" if (folder / "README.md").exists() else "")
    return ("| `final/audit/` | the whole residents layer | %d | — | — | — | %s | %d rows "
            "carry a research ticket |" % (len(rows), held, reviewed))


def index_block() -> str:
    note = (
        "Generated by `python3 tools/export_resident_research_package.py --index`; "
        "`--check` fails if it has drifted. Outcome spellings differ between the "
        "earlier one-off packagers and are normalized here (`corroborated` and "
        "`corroborated_enrichment` are the same outcome, as are `candidate` and "
        "`candidate_identity`); a `—` means that package's CSV records no outcome "
        "column, not that the count is zero. Row counts are CSV rows, so T-0493's "
        "345 are voter-list readings rather than 345 people. The rows are read off "
        "the directory, so a cohort package that lands without the index being "
        "rebuilt fails `--check` rather than going unlisted. The last row is not a "
        "cohort: `final/audit/` is the whole residents layer, one row per person, "
        "generated by `tools/export_resident_audit.py`."
    )
    return "\n".join([INDEX_START, "", note, ""] + index_rows() + ["", INDEX_END])


def write_index() -> int:
    path = REFERENCE / "README.md"
    text = path.read_text(encoding="utf-8")
    block = index_block()
    if INDEX_START in text and INDEX_END in text:
        head, rest = text.split(INDEX_START, 1)
        _, tail = rest.split(INDEX_END, 1)
        text = head + block + tail
    else:
        raise SystemExit("the reference README has no index:start/index:end markers")
    path.write_text(text, encoding="utf-8")
    print("index written: %d packages" % len(indexed_packages()))
    return 0


def index_is_current() -> list[str]:
    path = REFERENCE / "README.md"
    text = path.read_text(encoding="utf-8")
    if INDEX_START not in text or INDEX_END not in text:
        return ["the reference README has lost its index markers"]
    got = INDEX_START + text.split(INDEX_START, 1)[1].split(INDEX_END, 1)[0] + INDEX_END
    if got != index_block():
        return ["the reference README's package index has drifted — rebuild with --index"]
    return []


def csv_text(header: list[str], rows: list) -> str:
    import io
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\n")
    writer.writerow(header)
    for row in rows:
        writer.writerow(row)
    return buf.getvalue()


def residents_csv_text(rows: list[dict]) -> str:
    return csv_text(HEADERS, [[row[h] for h in HEADERS] for row in rows])


def write_workbook(path: Path, rows, candidate_rows, source_rows, search_rows, summary) -> bool:
    try:
        import openpyxl
    except ImportError:
        return False
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Residents"
    ws.append(HEADERS)
    for row in rows:
        ws.append([row[h] for h in HEADERS])
    for title, header, table in (
        ("Candidates", CANDIDATE_HEADER, candidate_rows),
        ("Sources", SOURCE_HEADER, source_rows),
        ("Search_Log", SEARCH_HEADER, search_rows),
    ):
        sheet = wb.create_sheet(title)
        sheet.append(header)
        for row in table:
            sheet.append(row)
    sheet = wb.create_sheet("Summary")
    sheet.append(["field", "value"])
    for key in ("ticket", "label", "size", "candidates", "sources", "searches"):
        sheet.append([key, summary[key]])
    for outcome in ("corroborated_enrichment", "candidate_identity", "no_corroboration"):
        sheet.append([outcome, summary["counts"].get(outcome, 0)])
    sheet.append(["reviewed_on", ", ".join(summary["reviewed_on"])])
    sheet.append(["derived_by", "tools/export_resident_research_package.py"])
    wb.save(path)
    return True


def package_files(ticket: str) -> dict:
    rows, candidate_rows, source_rows, search_rows, summary = build_rows(ticket)
    return {
        "rows": rows, "candidates": candidate_rows, "sources": source_rows,
        "searches": search_rows, "summary": summary,
        "%s_resident_research.csv" % ticket: residents_csv_text(rows),
        "README.md": readme_text(ticket, summary),
    }


def build(ticket: str) -> int:
    out = REFERENCE / ticket
    out.mkdir(parents=True, exist_ok=True)
    built = package_files(ticket)
    csv_name = "%s_resident_research.csv" % ticket
    (out / csv_name).write_text(built[csv_name], encoding="utf-8", newline="")
    (out / "README.md").write_text(built["README.md"], encoding="utf-8")
    xlsx = out / ("%s_resident_research_working.xlsx" % ticket)
    wrote = write_workbook(xlsx, built["rows"], built["candidates"], built["sources"],
                           built["searches"], built["summary"])
    summary = built["summary"]
    print("%s: %d residents, %d candidates, %d sources, %d searches%s"
          % (ticket, summary["size"], summary["candidates"], summary["sources"],
             summary["searches"], "" if wrote else "  (openpyxl unavailable — workbook skipped)"))
    return 0


def check(tickets: list[str]) -> int:
    problems = []
    for ticket in PACKAGES_OWED:
        out = REFERENCE / ticket
        if not out.is_dir():
            problems.append("%s: no package folder — a completed cohort owes one" % ticket)
            continue
        for name in ("%s_resident_research.csv" % ticket, "README.md"):
            if not (out / name).exists():
                problems.append("%s: %s is missing from the package" % (ticket, name))

    for ticket in tickets:
        out = REFERENCE / ticket
        built = package_files(ticket)
        csv_name = "%s_resident_research.csv" % ticket
        for name in (csv_name, "README.md"):
            path = out / name
            if not path.exists():
                problems.append("%s: %s is missing" % (ticket, name))
                continue
            if path.read_text(encoding="utf-8") != built[name]:
                problems.append("%s: %s no longer matches the records it was derived from "
                                "— rebuild with --build" % (ticket, name))
        xlsx = out / ("%s_resident_research_working.xlsx" % ticket)
        if not xlsx.exists():
            problems.append("%s: the working workbook is missing" % ticket)
        else:
            problems.extend(check_workbook(xlsx, ticket, built))

    problems.extend(index_is_current())

    if problems:
        for line in problems:
            print("FAIL  " + line)
        return 1
    print("resident-research packages ok — %d folders present, %d indexed, "
          "%d re-derived from source, index current"
          % (len(PACKAGES_OWED), len(indexed_packages()), len(tickets)))
    return 0


def check_workbook(path: Path, ticket: str, built: dict) -> list[str]:
    """A workbook is a zip with a build clock in it, so it cannot be byte-compared.

    What is compared is what a reader would notice: the four contracted sheets, and
    a Residents sheet that carries the same people as the CSV beside it.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return ["%s: the workbook will not open (%s)" % (ticket, exc)]
    problems = []
    missing = {"Residents", "Candidates", "Sources", "Search_Log"} - set(wb.sheetnames)
    if missing:
        problems.append("%s: the workbook is missing %s" % (ticket, ", ".join(sorted(missing))))
    if "Residents" in wb.sheetnames:
        ws = wb["Residents"]
        got = [str(row[2]) for row in ws.iter_rows(min_row=2, max_col=3, values_only=True)
               if row[2] not in (None, "")]
        want = [row["person_id"] for row in built["rows"]]
        if got != want:
            problems.append("%s: the workbook's Residents sheet holds %d people, the CSV %d "
                            "— rebuild with --build" % (ticket, len(got), len(want)))
    wb.close()
    return problems


def self_test() -> int:
    """The assertions above, each proved to fire."""
    failures = []

    payload = load(PAYLOAD)
    bad = json.loads(json.dumps(payload))
    bad["passes"][0]["counts"]["corroborated_enrichment"] += 1
    try:
        reviews_for("T-0442", bad)
        failures.append("a pass whose counts disagree with its own reviews was accepted")
    except SystemExit:
        pass

    try:
        reviews_for("T-9999", payload)
        failures.append("a ticket with no pass in the payload was accepted")
    except SystemExit:
        pass

    # The slice is taken by offset, so a pass must never read its neighbour's people.
    seen = set()
    for ticket in COHORTS:
        ids = {r["person_id"] for r in reviews_for(ticket, payload)}
        if ids & seen:
            failures.append("%s overlaps an earlier pass: %s" % (ticket, sorted(ids & seen)[:3]))
        seen |= ids
        manifest = load(RESEARCH / COHORTS[ticket]["manifest"])
        if {m["person_id"] for m in manifest["people"]} != ids:
            failures.append("%s: manifest and payload slice name different people" % ticket)

    # A candidate id is stable across runs and does not collide between people.
    cand = {"name": "Augustus Garrett"}
    if candidate_id("garrett_a", cand) != candidate_id("garrett_a", dict(cand)):
        failures.append("candidate ids are not stable")
    if candidate_id("garrett_a", cand) == candidate_id("garrett_b", cand):
        failures.append("candidate ids collide between people")

    # Every package the completion rule owes is readable, and the index names it.
    block = index_block()
    for ticket in PACKAGES_OWED:
        if ("`%s/`" % ticket) not in block:
            failures.append("%s is owed a package and the index does not name it" % ticket)

    # A package that lost its CSV must read as missing rather than as zero rows.
    if read_package("T-9999") is not None:
        failures.append("a folder with no CSV was read as a package")

    if failures:
        for line in failures:
            print("FAIL  " + line)
        return 1
    print("export_resident_research_package self-test ok — %d assertions"
          % (5 + 2 * len(COHORTS) + len(PACKAGES_OWED)))
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("ticket", nargs="?", help="cohort ticket, e.g. T-0442")
    ap.add_argument("--all", action="store_true", help="every cohort this tool can export")
    ap.add_argument("--build", action="store_true", help="write the package")
    ap.add_argument("--index", action="store_true",
                    help="rewrite the package index in the reference README")
    ap.add_argument("--check", action="store_true", help="re-derive and compare (the gate)")
    ap.add_argument("--self-test", action="store_true", help="prove the assertions fire")
    args = ap.parse_args()

    if args.self_test:
        return self_test()

    if args.ticket:
        if args.ticket not in COHORTS:
            raise SystemExit("%s: this tool exports %s" % (args.ticket, ", ".join(COHORTS)))
        tickets = [args.ticket]
    else:
        tickets = list(COHORTS)

    if args.index:
        return write_index()

    if args.build:
        for ticket in tickets:
            build(ticket)
        return write_index()
    return check(tickets)


if __name__ == "__main__":
    sys.exit(main())
