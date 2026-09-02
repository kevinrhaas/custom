#!/usr/bin/env python3
"""One-time T-0490 repair runner. Removed by its workflow after use."""
from __future__ import annotations

import json
import subprocess
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
SYNTH = ROOT / "tools" / "synthesize_resident_research.py"
RESEARCH = ROOT / "data" / "research" / "residents"
CENSUS = ROOT.parent.parent / "reference" / "census1840" / "validation"


def replace_once(text: str, old: str, new: str, label: str) -> str:
    if old not in text:
        raise SystemExit(f"{label} source pattern not found")
    return text.replace(old, new, 1)


def patch_synthesis() -> None:
    s = SYNTH.read_text(encoding="utf-8")
    s = replace_once(
        s,
        '''                hh["arrival"] = {"value": f"{year:04d}-01-01", "confidence": "attested",
                    "sources": srcs, "precision": "year",
                    "note": f"YEAR PRECISION ONLY. {item.get('ticket')} states arrival/move to Chicago in {year}; January 1 is only the dataset's year anchor."}''',
        '''                hh["arrival"] = {"value": f"{year:04d}", "confidence": "attested",
                    "sources": srcs, "precision": "year",
                    "note": f"YEAR PRECISION ONLY. {item.get('ticket')} states arrival/move to Chicago in {year}; no month or day is asserted."}''',
        "arrival",
    )
    s = replace_once(
        s,
        '''            if not (p.get("note") or "").startswith(prefix): p["note"]=prefix+(p.get("note") or "")''',
        '''            existing = p.get("note") or ""
            existing = re.sub(r"^(?:INDEPENDENTLY CORROBORATED RESIDENT\\. |PROJECTED RESIDENT\\. Documented in Chicago post-office evidence but not independently corroborated strongly enough for attested circa-1835 residence\\. )", "", existing, flags=re.I)
            if outcome in CORROBORATED:
                existing = re.sub(r"^KNOWN ONLY FROM THE POST OFFICE\\.\\s*", "", existing, flags=re.I)
                existing = re.sub(r"Nothing else in the corpus names this person[^.]*\\.\\s*", "", existing, flags=re.I)
                existing = re.sub(r"No (?:arrival|trade|occupation)[^.]*\\.\\s*", "", existing, flags=re.I)
                prefix = "INDEPENDENTLY CORROBORATED RESIDENT. Originally documented in Chicago post-office evidence; independent resident research now corroborates the identity. "
            p["note"]=(prefix+existing).strip()''',
        "resident note",
    )
    s = replace_once(
        s,
        '''    index=load(INDEX); before=snapshot(index); docs={p:load(p) for p in sorted(HOUSEHOLDS.glob("*.json"))}; research=research_rows()''',
        '''    index=load(INDEX); current_before=snapshot(index)
    prior_ledger=load(LEDGER) if LEDGER.exists() else {}
    before=(prior_ledger.get("before") if current_before.get("reconstructed")==0 and prior_ledger.get("before") else current_before)
    docs={p:load(p) for p in sorted(HOUSEHOLDS.glob("*.json"))}; research=research_rows()''',
        "baseline",
    )
    s = replace_once(
        s,
        '''    persons={p.get("id"):(p,d) for d in docs.values() for p in d.get("persons") or [] if p.get("id")}''',
        '''    if prior_ledger and stats["removed_people"] == 0 and stats["removed_households"] == 0:
        prior_retirement=prior_ledger.get("retirement") or {}
        stats["removed_people"]=int(prior_retirement.get("removed_people") or 0)
        stats["removed_households"]=int(prior_retirement.get("removed_households") or 0)
    persons={p.get("id"):(p,d) for d in docs.values() for p in d.get("persons") or [] if p.get("id")}''',
        "retirement",
    )
    SYNTH.write_text(s, encoding="utf-8")


def export_census_bridge() -> None:
    from openpyxl import load_workbook

    wanted = {
        "1835 Tier", "Tier Confidence", "Include Recommendation", "Preferred Name",
        "1835 Poll", "1834 Poll", "1833 Poll", "1833 Tax", "1840 Census Page", "1840 Row",
        "IPUMS SERIAL / Candidate Block", "1840 Name Confidence", "1839 directory match",
        "Reasoning", "reasoning", "Notes", "notes",
    }
    rows = []
    for path in sorted(CENSUS.glob("Chicago_1835_Best_Resident_Set_Research*.xlsx")):
        wb = load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            headers = None
            header_row = None
            for rno, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 30), values_only=True), 1):
                vals = [str(v).strip() if v is not None else "" for v in row]
                if "Preferred Name" in vals and ("1835 Tier" in vals or "Include Recommendation" in vals):
                    headers, header_row = vals, rno
                    break
            if not headers:
                continue
            for rno, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
                rec = {headers[i]: row[i] for i in range(min(len(headers), len(row))) if headers[i]}
                if not rec.get("Preferred Name"):
                    continue
                selected = {"file": path.name, "sheet": ws.title, "row": rno}
                for key, val in rec.items():
                    low = key.lower()
                    if key in wanted or any(x in low for x in ("reason", "recommend", "candidate", "alternate", "raw reading")):
                        if val not in (None, ""):
                            selected[key] = val
                rows.append(selected)
        wb.close()
    (RESEARCH / "census_1835_bridge_candidates.json").write_text(
        json.dumps({"generated": "2026-09-02", "rows": rows}, indent=2, ensure_ascii=False, default=str) + "\n",
        encoding="utf-8",
    )
    print(f"exported {len(rows)} census bridge rows")


def capture(command: list[str], filename: str) -> int:
    proc = subprocess.run(command, cwd=ROOT, text=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
    (RESEARCH / filename).write_text(proc.stdout, encoding="utf-8")
    print(f"{' '.join(command)} => {proc.returncode}")
    return proc.returncode


def main() -> int:
    patch_synthesis()
    subprocess.run(["python", "tools/synthesize_resident_research.py"], cwd=ROOT, check=True)
    subprocess.run(["python", "tools/synthesize_resident_research.py", "--check"], cwd=ROOT, check=True)
    export_census_bridge()
    capture(["python", "tools/validate.py", "--all"], "synthesis_validate_after_fixes.log")
    capture(["bash", "tools/check.sh"], "synthesis_full_gate_after_fixes.log")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
