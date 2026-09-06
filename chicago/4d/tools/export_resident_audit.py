#!/usr/bin/env python3
"""The final audit of the resident-research programme: one row per person, and what it rests on.

WHY THIS EXISTS, in the owner's words (2026-09-03): "I am concerned that there are only
adjudicated mappings through v4 and none other and we did at least 11 slices of residents
to get to households but most of your census work needs to be published". Asked what
"published" should deliver, he chose "Reference packages + final audit".

T-0490's acceptance clause promised `chicago/reference/resident-research/final/audit/`
with a master workbook, a CSV and a README carrying coverage metrics and a gaps list. That
ticket closed (PR #668) and none of it existed on `dev`. T-0512 is that promise, kept.

WHAT IT IS. A DERIVED read of the committed residents layer and the research ledgers
beside it, written out as a flat table a person can open in a spreadsheet and sort. One row
per person in `data/residents/households/*.json`: who they are, which research ticket
looked at them and what it concluded, which source ids stand behind them BY CATEGORY, the
unresolved flags, and a one-word audit result that says what the person rests on.

WHAT IT DELIBERATELY IS NOT. It alters no research outcome, mints nobody, and grades
nothing. Every column is copied or counted from a committed record; where this tool has to
JUDGE — and it judges in exactly one place, which category a source id belongs to — the
judgement is a written table below, printed into the README, and total: a source id this
tool has never met stops the build rather than falling into a bucket unseen.

    tools/export_resident_audit.py --build       write the package
    tools/export_resident_audit.py --check       the committed package still re-derives
    tools/export_resident_audit.py --self-test   the assertions still fire when broken
    tools/export_resident_audit.py --report      the metrics table, to stdout

THE XLSX IS OPTIONAL AND THE CSV IS NOT. `openpyxl` is not in every sandbox this repo is
worked from, so the CSV and the README are always written and always gated; the workbook is
written when openpyxl imports and is never compared byte for byte (a zip carries its own
timestamps). `--check` fails on a CSV or README that has drifted from the layer — a hand
edit to the audit is refused the same way a hand edit to any generated file here is.
"""
from __future__ import annotations

import argparse
import csv
import io
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]                 # chicago/4d
REFERENCE = ROOT.parent / "reference"                      # chicago/reference
OUT = REFERENCE / "resident-research" / "final" / "audit"
HOUSEHOLDS = ROOT / "data/residents/households"
INDEX = ROOT / "data/residents/index.json"
RESEARCH = ROOT / "data/research/residents"
RULINGS = RESEARCH / "conflict_rulings.json"
SOURCES = ROOT / "data/sources"

CSV_NAME = "resident_audit_master.csv"
XLSX_NAME = "resident_audit_master.xlsx"
README_NAME = "README.md"

# ---------------------------------------------------------------- the one judgement
#
# THE SOURCE CATEGORY TABLE. The ticket asks for source coverage by category — newspaper,
# civic, census, church, book, directory, secondary — and the source records carry a
# `type` that is a MEDIA class (book, website, dataset, newspaper, map, legal…), not that.
# So the mapping is written out here rather than sniffed, in the order the rules fire, and
# `--check` refuses a source id no rule reaches. A reader who disagrees with a placement
# can argue with a line of this table; nothing here is hidden behind a heuristic.
#
# Rules 1-4 are by IDENTITY of the record and beat the media class. Rules 5-7 fall back to
# the source record's own `type`, which is where the long tail of county histories and web
# pages lands.
CENSUS_IDS = {
    "census_1840_chicago_familysearch_images",
    "census_1840_chicago_v4_research",
    "dalton_1840_chicago_census_index",
    "resident_research_v4_1835_census_bridge",
}
CHURCH_IDS = {
    "baptisthistoryhomepage",
    "catholic_chicago_st_cyr_1833",
    "first_presbyterian_chicago_1833_1913",
    "resident_research_schindler_first_church",
    "rr_baptist_chicago_lathrop",
    "st_cyr_register_ichr_v4",
}
# Civic: a record made by a government or a public body about a named man — poll books and
# voter lists, tax and muster rolls, treaty payment schedules, land patents, coroner's
# returns, congressional and law-report biographies of officeholders.
CIVIC_IDS = {
    "blackhawk_war_chicago_enrollments_isa",
    "chicago_voter_lists_1833_1835_irad",
    "fcp_treaty_chicago_1833_payments",
    "ida_will_property_1842_hawley",
    "isa_public_domain_land_tract_sales",
    "lake_county_coroner_starr_titus_1839",
    "okstate_treaty_chicago_1833",
    "resident_research_burnap_jersey_land",
    "resident_research_lincoln_walker_1835",
    "resident_research_meacham_uscg",
    "resident_research_usreports_wooley",
    "rr_cook_democratic_republicans_1834",
    "ushouse_james_h_woodworth",
    "ushouse_john_reynolds_illinois",
}
DIRECTORY_RE = re.compile(r"directory")
TYPE_TO_CATEGORY = {
    "newspaper": "newspaper",
    "book": "book",
    "article": "book",
    "manuscript": "book",
    "website": "secondary",
    "dataset": "secondary",
    "map": "secondary",
    "legal": "secondary",
    "illustration": "secondary",
    "photograph": "secondary",
}
CATEGORIES = ["newspaper", "civic", "census", "church", "book", "directory", "secondary"]

KIN = {"wife", "husband", "son", "daughter", "child", "brother", "sister",
       "mother", "father"}

AUDIT_RESULTS = [
    "corroborated_across_categories",
    "two_or_more_sources_one_category",
    "one_source",
    "the_letter_lists_alone",
    "no_source",
]



class Refused(Exception):
    """The build met something it will not guess at."""


# ---------------------------------------------------------------- reading the layer

def source_type(source_id: str, cache: dict) -> str | None:
    if source_id not in cache:
        path = SOURCES / ("%s.json" % source_id)
        cache[source_id] = json.loads(path.read_text())["type"] if path.exists() else None
    return cache[source_id]


def category_of(source_id: str, cache: dict) -> str:
    if source_id in CENSUS_IDS:
        return "census"
    if DIRECTORY_RE.search(source_id):
        return "directory"
    if source_id in CHURCH_IDS:
        return "church"
    if source_id in CIVIC_IDS:
        return "civic"
    kind = source_type(source_id, cache)
    if kind is None:
        raise Refused(
            "source id %r is cited by the residents layer and has no record in "
            "data/sources/ — categorise it there before the audit can count it" % source_id)
    if kind not in TYPE_TO_CATEGORY:
        raise Refused(
            "source %r has type %r, which the audit's category table does not know. Add a "
            "rule rather than letting it fall into a bucket unseen." % (source_id, kind))
    return TYPE_TO_CATEGORY[kind]


def households() -> list[dict]:
    return [json.loads(p.read_text()) for p in sorted(HOUSEHOLDS.glob("*.json"))]


def findings() -> dict:
    """person id -> the pass ledger's override for that person, where one was written.

    The card carries the OUTCOME; the ledger carries the candidates and the conflicts that
    produced it, and a conflict nobody can see is the one an audit exists to surface."""
    out = {}
    for path in sorted(RESEARCH.glob("pass_*_findings.json")):
        doc = json.loads(path.read_text())
        for person_id, override in (doc.get("overrides") or {}).items():
            row = dict(override)
            row["_ticket"] = doc.get("ticket")
            out[person_id] = row
    return out


def ledger_conflicts() -> dict[str, list[str]]:
    """person id -> every conflict string any pass recorded against them, in pass order.

    NOT `findings()`, and that distinction is the whole of T-0845. `findings()` ASSIGNS
    (`out[person_id] = row`), so a person two passes reviewed keeps only the last override —
    which is right for the OUTCOME a card carries and wrong for the conflicts, because a
    conflict pass 02 wrote is still written after pass 13 rewrites the row around it. Read
    the newest override only and the audit reported conflicts against 68 people where the
    ledgers hold them against 96; the twenty-eight in the gap were not resolved, they were
    overwritten, and T-0733 never ruled on them because nothing could see them."""
    out: dict[str, list[str]] = {}
    for path in sorted(RESEARCH.glob("pass_*_findings.json")):
        doc = json.loads(path.read_text())
        for person_id, override in (doc.get("overrides") or {}).items():
            for candidate in (override.get("candidates") or []):
                for conflict in (candidate.get("conflicts") or []):
                    out.setdefault(person_id, []).append(conflict)
    return out


def ruling_covers(ruling: dict, recorded_conflicts: list[str]) -> bool:
    """Does this ruling reach the conflicts the ledgers state RIGHT NOW?

    Exact-list equality, deliberately. A ruling names the conflict strings it was made
    against; anything else on the record is a conflict it never read."""
    if not ruling:
        return False
    return list(ruling.get("ruled_conflicts") or []) == list(recorded_conflicts)


def rulings() -> dict:
    """The written adjudications of the conflicts the ledgers record — T-0733.

    A ruling is PINNED to the conflict text it read. `ruled_conflicts` carries the
    ledger's conflict strings verbatim, and the ruling covers a person only while that
    list is still exactly what the ledgers say about them. Reword a conflict, or add a
    second one to a person already ruled, and the ruling stops reaching them — the flag
    fires again and somebody has to look. A ruling that could rubber-stamp a conflict it
    has never read would be worse than no ruling at all."""
    doc = json.loads(RULINGS.read_text())
    for verdict in sorted({r["verdict"] for r in doc["rulings"].values()}):
        if verdict not in doc["verdicts"]:
            raise Refused(
                "conflict_rulings.json rules with verdict %r and does not define it. "
                "Every verdict states what it means and what it does NOT mean, or it is "
                "not a ruling." % verdict)
    for person_id, row in doc["rulings"].items():
        if not row.get("reopens_on"):
            raise Refused(
                "the ruling on %r names no reopening condition. A decline that cannot be "
                "reopened is a closed door, and none of these are." % person_id)
    return doc


def value_of(block) -> str:
    """A graded block's value as one cell. `{value, confidence, note}` is the shape of
    nearly every claim in this layer; a bare scalar is passed through."""
    if isinstance(block, dict):
        v = block.get("value")
        return "" if v is None else str(v)
    return "" if block is None else str(block)


def cited_sources(person: dict, household: dict) -> list[str]:
    """Every source id that stands behind this person, from every place one is written.

    The card cites at three depths — the person's own `sources`, the evidence blocks the
    consolidation wrote, and the graded household blocks the mint wrote — and a person
    whose only census link is inside `later_census` is exactly the person an audit that
    read only the top level would report as uncensused."""
    ids: list[str] = list(person.get("sources") or [])
    ids += (person.get("resident_research") or {}).get("source_ids") or []
    for block in ("press_evidence", "civic_evidence", "book_evidence",
                  "church_evidence", "census_evidence"):
        for entry in person.get(block) or []:
            if entry.get("source"):
                ids.append(entry["source"])
    later = person.get("later_census")
    if isinstance(later, dict) and later.get("source_id"):
        ids.append(later["source_id"])
    occupation = person.get("occupation")
    if isinstance(occupation, dict):
        ids += occupation.get("sources") or []
    if person.get("relationship") == "head":
        for key in ("arrival", "origin", "lives_at", "works_at",
                    "present_on_scene_date", "party_size_on_arrival", "reason_for_coming"):
            block = household.get(key)
            if isinstance(block, dict):
                ids += block.get("sources") or []
    seen, out = set(), []
    for source_id in ids:
        if source_id and source_id not in seen:
            seen.add(source_id)
            out.append(source_id)
    return out


LETTER_LIST_SOURCES = {"chicago_democrat_1833_1835", "chicago_american_1835",
                       "chicago_democrat_1833_11_26"}

COLUMNS = [
    "person_id", "household_id", "name", "relationship", "grade", "ladder_rule",
    "resident_subtype", "letter_list_only", "letter_list_returns",
    "letter_list_return_dates", "research_ticket", "research_outcome",
    "research_reviewed_on", "asserted_identity", "sources_total", "source_ids",
    "categories_covered", "src_newspaper", "src_civic", "src_census", "src_church",
    "src_book", "src_directory", "src_secondary", "later_census_year",
    "later_census_serial", "later_census_bridge_status", "present_on_scene_date",
    "division", "lives_at", "works_at", "occupation", "occupation_confidence",
    "conflict_recorded", "conflict_ruling", "conflict_reopens_on",
    "flag_no_research_row", "flag_candidate_identity_open", "flag_conflicting_evidence",
    "flag_standing_constraint",
    "flag_single_source", "flag_no_source", "flag_unplaced", "flag_no_address",
    "audit_result",
]


def rows() -> tuple[list[dict], dict]:
    ledger = findings()
    conflicts_recorded = ledger_conflicts()
    ruled = rulings()
    seen_rulings: set[str] = set()
    live_households: set[str] = set()
    cache: dict = {}
    out: list[dict] = []
    for household in households():
        live_households.add(household["id"])
        members = household.get("persons") or []
        for person in members:
            ids = cited_sources(person, household)
            by_category = defaultdict(list)
            for source_id in ids:
                by_category[category_of(source_id, cache)].append(source_id)
            research = person.get("resident_research") or {}
            has_research_row = bool(research.get("ticket"))
            override = ledger.get(person["id"]) or {}
            recorded_conflicts = conflicts_recorded.get(person["id"], [])
            ruling = ruled["rulings"].get(person["id"]) or {}
            if ruling:
                seen_rulings.add(person["id"])
            # The ruling reaches this person only while it still names exactly the
            # conflicts the ledgers state. See rulings().
            covered = ruling_covers(ruling, recorded_conflicts)
            standing = bool(household.get("review_required"))
            later = person.get("later_census") if isinstance(
                person.get("later_census"), dict) else {}
            returns = person.get("letter_list_returns") or []
            categories_covered = sum(1 for c in CATEGORIES if by_category[c])
            only_letter_lists = (
                bool(person.get("letter_list_only"))
                and set(ids) <= LETTER_LIST_SOURCES)
            if not ids:
                result = "no_source"
            elif only_letter_lists:
                result = "the_letter_lists_alone"
            elif len(ids) == 1:
                result = "one_source"
            elif categories_covered >= 2:
                result = "corroborated_across_categories"
            else:
                result = "two_or_more_sources_one_category"
            row = {
                "person_id": person["id"],
                "household_id": household["id"],
                "name": person.get("name") or "",
                "relationship": person.get("relationship") or "",
                "grade": person.get("grade") or "",
                "ladder_rule": person.get("ladder_rule") or "",
                "resident_subtype": person.get("resident_subtype") or "",
                "letter_list_only": bool(person.get("letter_list_only")),
                "letter_list_returns": len(returns),
                "letter_list_return_dates": ";".join(returns),
                "research_ticket": research.get("ticket") or "",
                "research_outcome": research.get("outcome") or "",
                "research_reviewed_on": research.get("reviewed_on") or "",
                "asserted_identity": research.get("asserted_identity", ""),
                "sources_total": len(ids),
                "source_ids": ";".join(ids),
                "categories_covered": categories_covered,
                "later_census_year": later.get("year", ""),
                "later_census_serial": later.get("serial", ""),
                "later_census_bridge_status": later.get("bridge_status", ""),
                "present_on_scene_date": value_of(household.get("present_on_scene_date")),
                "division": household.get("division") or "",
                "lives_at": value_of(household.get("lives_at")),
                "works_at": value_of(household.get("works_at")),
                "occupation": value_of(person.get("occupation")),
                "occupation_confidence": (person.get("occupation") or {}).get(
                    "confidence", "") if isinstance(person.get("occupation"), dict) else "",
                "flag_no_research_row": not has_research_row,
                "flag_candidate_identity_open": (
                    research.get("outcome") in ("candidate_identity", "candidate")
                    and not research.get("asserted_identity")),
                "conflict_recorded": bool(recorded_conflicts),
                "conflict_ruling": ruling.get("verdict", "") if covered else "",
                "conflict_reopens_on": ruling.get("reopens_on", "") if covered else "",
                "flag_standing_constraint": standing,
                "flag_conflicting_evidence": bool(recorded_conflicts) and not covered,
                "flag_single_source": len(ids) == 1,
                "flag_no_source": not ids,
                "flag_unplaced": household.get("division") == "unplaced",
                "flag_no_address": not value_of(household.get("lives_at"))
                and not value_of(household.get("works_at")),
                "audit_result": result,
            }
            for category in CATEGORIES:
                row["src_%s" % category] = ";".join(by_category[category])
            out.append(row)
    out.sort(key=lambda r: (r["household_id"], r["person_id"]))
    orphans = sorted(set(ruled["rulings"]) - seen_rulings)
    if orphans:
        raise Refused(
            "conflict_rulings.json rules on %d person(s) the residents layer no longer "
            "holds: %s. A ruling outlives its person only when a card was merged or "
            "renamed under it — move the ruling or retire it."
            % (len(orphans), ", ".join(orphans)))
    stray = sorted(set(ruled["standing_constraints"]) - live_households)
    if stray:
        raise Refused(
            "conflict_rulings.json names standing constraints on household(s) that are "
            "gone: %s" % ", ".join(stray))
    return out, cache


# ---------------------------------------------------------------- the metrics

def metrics(table: list[dict]) -> dict:
    total = len(table)

    def n(predicate) -> int:
        return sum(1 for r in table if predicate(r))

    sizes = Counter(r["household_id"] for r in table)
    return {
        "persons": total,
        "identities": n(lambda r: r["sources_total"] > 0),
        "occupations": n(lambda r: r["occupation"] not in ("", "none_recorded")),
        "household_membership": n(lambda r: sizes[r["household_id"]] > 1),
        "kinship": n(lambda r: r["relationship"] in KIN),
        "property_address": n(lambda r: r["lives_at"] or r["works_at"]),
        "voter_civic_evidence": n(lambda r: r["src_civic"]),
        "census_linkage": n(lambda r: r["src_census"] or r["later_census_year"] != ""),
    }


def gaps(table: list[dict]) -> list[tuple[str, int, str]]:
    """The named list of what is still open, largest first. Each row is a gap a later
    ticket can be written against, not a defect in this export."""
    total = len(table)
    named = [
        ("no research row", sum(1 for r in table if r["flag_no_research_row"]),
         "no cohort ticket has reviewed this person; the programme reached %d of %d"
         % (total - sum(1 for r in table if r["flag_no_research_row"]), total)),
        ("rests on one source", sum(1 for r in table if r["flag_single_source"]),
         "one source id on the card and no second category to check it against"),
        ("rests on the letter lists alone",
         sum(1 for r in table if r["audit_result"] == "the_letter_lists_alone"),
         "known only from the post office's uncalled-for lists"),
        ("unplaced", sum(1 for r in table if r["flag_unplaced"]),
         "the household carries division `unplaced`: in the town, on no lot"),
        ("no address", sum(1 for r in table if r["flag_no_address"]),
         "neither `lives_at` nor `works_at` resolves"),
        ("candidate identity open",
         sum(1 for r in table if r["flag_candidate_identity_open"]),
         "a candidate was found and not asserted; the identity is still a question"),
        ("conflicting evidence, unruled",
         sum(1 for r in table if r["flag_conflicting_evidence"]),
         "the ledger records a conflict against a candidate and no ruling in "
         "`data/research/residents/conflict_rulings.json` reaches it (T-0733)"),
        ("conflicting evidence, ruled",
         sum(1 for r in table if r["conflict_ruling"]),
         "a recorded conflict carries a written adjudication and a named reopening "
         "condition; every one of them is a decline, and none adopts a candidate"),
        ("standing constraint",
         sum(1 for r in table if r["flag_standing_constraint"]),
         "the household carries `review_required` with `touches_removal`: the final "
         "removal of the Potawatomi reaches it, no scene holding it may be `released`, "
         "and no research retires the flag"),
        ("no census linkage",
         sum(1 for r in table if not (r["src_census"] or r["later_census_year"] != "")),
         "no 1840 census row is bridged to this person"),
        ("no source of their own",
         sum(1 for r in table if r["flag_no_source"]),
         "the collective `household_member` rows — \"the rest of the Beaubien "
         "household, unnamed\" and its two fellows — which are an inferred count of "
         "people, not named individuals; the household record carries the sources"),
    ]
    return sorted(named, key=lambda g: -g[1])


# ---------------------------------------------------------------- writing

def render_csv(table: list[dict]) -> str:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=COLUMNS, lineterminator="\n")
    writer.writeheader()
    for row in table:
        writer.writerow({k: row[k] for k in COLUMNS})
    return buf.getvalue()


def render_readme(table: list[dict], cache: dict) -> str:
    m = metrics(table)
    total = m["persons"]
    results = Counter(r["audit_result"] for r in table)
    tickets = Counter(r["research_ticket"] for r in table if r["research_ticket"])
    outcomes = Counter(r["research_outcome"] for r in table if r["research_outcome"])
    categories = Counter()
    for row in table:
        for category in CATEGORIES:
            if row["src_%s" % category]:
                categories[category] += 1

    lines = []
    add = lines.append
    add("# The final audit of the resident-research programme")
    add("")
    add("GENERATED by `chicago/4d/tools/export_resident_audit.py --build`. Hand-edit and")
    add("`--check` says so — it runs in `tools/check.sh` and rebuilds this package from")
    add("`chicago/4d/data/residents/` on every commit.")
    add("")
    add("This is what T-0490's acceptance clause promised and never produced, delivered")
    add("under **T-0512**: one row per person in the town's residents layer, saying which")
    add("research ticket looked at them, what it concluded, which source ids stand behind")
    add("them by category, what is still unresolved, and — in one word — what the person")
    add("rests on.")
    add("")
    add("| file | what it is |")
    add("| --- | --- |")
    add("| `%s` | the table, one row per person, %d rows and %d columns |"
        % (CSV_NAME, total, len(COLUMNS)))
    add("| `%s` | the same table as a workbook, plus the metrics, gaps and category "
        "sheets. Written when `openpyxl` imports; the CSV is the gated artifact |"
        % XLSX_NAME)
    add("| `%s` | this file |" % README_NAME)
    add("")
    add("**It changes nothing.** No research outcome is altered here, nobody is minted,")
    add("no grade moves. Every cell is copied or counted from a committed record.")
    add("")
    add("## Coverage")
    add("")
    add("Each line is *how many of the %d people carry at least one record of that kind*."
        % total)
    add("")
    add("| coverage | of %d | %% | what counts |" % total)
    add("| --- | ---: | ---: | --- |")
    labels = {
        "identities": "identities",
        "occupations": "occupations",
        "household_membership": "household membership",
        "kinship": "kinship",
        "property_address": "property / address",
        "voter_civic_evidence": "voter / civic evidence",
        "census_linkage": "census linkage",
    }
    whats = {
        "identities": "a name with at least one source id anywhere on the card",
        "occupations": "an occupation that is not `none_recorded`",
        "household_membership": "recorded inside a household of two or more people",
        "kinship": "a stated kin relationship (%s)" % ", ".join(sorted(KIN)),
        "property_address": "the household resolves a `lives_at` or a `works_at`",
        "voter_civic_evidence": "a poll book, tax list, muster roll, treaty payment or "
                                "other public record",
        "census_linkage": "an 1840 census row bridged to this person",
    }
    for key in ("identities", "occupations", "household_membership", "kinship",
                "property_address", "voter_civic_evidence", "census_linkage"):
        add("| %s | **%d** | %.1f%% | %s |"
            % (labels[key], m[key], 100.0 * m[key] / total, whats[key]))
    add("")
    add("## What each person rests on")
    add("")
    add("| audit result | people | % |")
    add("| --- | ---: | ---: |")
    for key in AUDIT_RESULTS:
        add("| `%s` | %d | %.1f%% |"
            % (key, results.get(key, 0), 100.0 * results.get(key, 0) / total))
    add("")
    add("`corroborated_across_categories` is the only result that means two *kinds* of")
    add("record agree; two newspaper notices of the same name are")
    add("`two_or_more_sources_one_category` and no stronger.")
    add("")
    add("## Source coverage by category")
    add("")
    add("| category | people citing at least one |")
    add("| --- | ---: |")
    for category in CATEGORIES:
        add("| %s | %d |" % (category, categories.get(category, 0)))
    add("")
    add("The category of a source id is the audit's one judgement, and it is a written")
    add("table in the tool rather than a heuristic: census, directory, church and civic")
    add("are named record by record; everything else falls back to the source record's own")
    add("`type` (`newspaper` → newspaper, `book`/`article`/`manuscript` → book, the rest →")
    add("secondary). A source id no rule reaches stops the build.")
    add("")
    add("| source id | category | type |")
    add("| --- | --- | --- |")
    for source_id in sorted(cache):
        add("| `%s` | %s | %s |"
            % (source_id, category_of(source_id, cache), source_type(source_id, cache)))
    add("")
    add("## The research programme")
    add("")
    add("| cohort ticket | people reviewed |")
    add("| --- | ---: |")
    for ticket, count in sorted(tickets.items()):
        add("| %s | %d |" % (ticket, count))
    add("| **reviewed** | **%d** |" % sum(tickets.values()))
    add("| **not yet reviewed** | **%d** |" % (total - sum(tickets.values())))
    add("")
    add("| research outcome | people |")
    add("| --- | ---: |")
    for outcome, count in sorted(outcomes.items(), key=lambda kv: -kv[1]):
        add("| `%s` | %d |" % (outcome, count))
    add("")
    add("## The conflicts, and what was ruled on them")
    add("")
    add("Under **T-0733**. The ledgers record a conflict against a candidate for **%d**"
        % sum(1 for r in table if r["conflict_recorded"]))
    add("of the %d people. Before T-0733 nothing ruled on any of them, and a conflict" % total)
    add("that is recorded and never adjudicated reads, to anybody downstream, exactly like")
    add("a conflict nobody found. `data/research/residents/conflict_rulings.json` is the")
    add("adjudication: a verdict, the conflict text it was made against, and the record")
    add("that would reopen it.")
    add("")
    add("**Every verdict is a decline, and no candidate is adopted here.** The decline is")
    add("what the layer already did silently — the candidate was never asserted and the")
    add("card carries no identity on its strength. What was missing was the writing down.")
    add("")
    add("| verdict | people | what it means |")
    add("| --- | ---: | --- |")
    verdicts = Counter(r["conflict_ruling"] for r in table if r["conflict_ruling"])
    ruled_doc = rulings()
    for verdict, count in sorted(verdicts.items(), key=lambda kv: (-kv[1], kv[0])):
        add("| `%s` | %d | %s |"
            % (verdict, count, " ".join(ruled_doc["verdicts"][verdict].split())))
    add("| **unruled** | **%d** | a recorded conflict no ruling reaches; the flag fires |"
        % sum(1 for r in table if r["flag_conflicting_evidence"]))
    add("")
    add("A ruling is pinned to the conflict text it read: `ruled_conflicts` carries the")
    add("ledger's strings verbatim, and the moment one is reworded — or a second conflict")
    add("is added to a person already ruled — the ruling stops reaching them and")
    add("`flag_conflicting_evidence` fires again. A ruling may not rubber-stamp a conflict")
    add("it has never read.")
    add("")
    add("### Every pass, not the newest one")
    add("")
    add("**T-0845.** A person's conflicts are read from EVERY pass that reviewed them.")
    add("`findings()` keeps only the last override written for a person, which is right for")
    add("the outcome a card carries and wrong for the conflicts: one written in pass 02 is")
    add("still written after pass 13 rewrites the row around it. Read the newest override")
    add("alone and this audit reported conflicts against 68 people where the ledgers hold")
    add("them against %d. The twenty-eight in the gap were not resolved — they were"
        % sum(1 for r in table if r["conflict_recorded"]))
    add("overwritten, and T-0733 ruled on none of them because nothing could see them. The")
    add("worst was Angeline Vann, whose conflict is the one in the set that DISQUALIFIES")
    add("rather than fails to bridge: she was born in 1834, and an infant is not the person")
    add("a letter waits for. She is ruled `refused_date_excludes` — the only verdict here")
    add("that is not a decline, because a decline says the bridge was not found and this")
    add("one says no such bridge can exist.")
    add("")
    add("### The standing constraints are not conflicts")
    add("")
    add("%d people in %d households carry `review_required`, and folding them into the"
        % (sum(1 for r in table if r["flag_standing_constraint"]),
           len(ruled_doc["standing_constraints"])))
    add("conflicting-evidence flag was a defect in the flag. Every one of those households")
    add("carries `touches_removal` with it: they are the households the final removal of")
    add("the Potawatomi from Chicago reaches, AGENTS.md requires the flag to stand on them")
    add("permanently, and `tools/measure_review_constraint.py --gate` refuses its silent")
    add("removal. It is a standing ethical constraint, not an open evidentiary question,")
    add("and no research retires it. They are counted under `flag_standing_constraint`.")
    add("")
    add("| household | people |")
    add("| --- | ---: |")
    for hid in sorted(ruled_doc["standing_constraints"]):
        add("| `%s` | %d |"
            % (hid, len(ruled_doc["standing_constraints"][hid]["persons"])))
    add("")
    add("## The gaps")
    add("")
    add("Named, largest first. Each is a ticket somebody could write; none of them is a")
    add("defect in this export.")
    add("")
    add("| gap | people | what it means |")
    add("| --- | ---: | --- |")
    for name, count, what in gaps(table):
        add("| %s | %d | %s |" % (name, count, what))
    add("")
    add("## Reading the table")
    add("")
    add("- `source_ids` is every source id behind the person, from the card's own")
    add("  `sources`, the research row, the five evidence blocks the consolidation wrote,")
    add("  the `later_census` bridge, the occupation block, and — for a head — the")
    add("  household's own graded blocks. `src_<category>` splits the same list.")
    add("- `conflict_recorded` says the ledgers hold a conflict against a candidate for")
    add("  this person; `conflict_ruling` and `conflict_reopens_on` are the adjudication")
    add("  T-0733 wrote, and are empty when no ruling reaches the conflicts on record.")
    add("- `flag_conflicting_evidence` means an UNRULED conflict — the flag fires on what")
    add("  nobody has looked at, not on the fact that a conflict exists. Before T-0733 it")
    add("  meant \"a conflict exists\" and also swept in the `review_required` households,")
    add("  which are a standing constraint and not a conflict at all; those now count")
    add("  under `flag_standing_constraint`. T-0517 compares against that definition.")
    add("- `flag_*` columns are the unresolved list. They are not failures; they are what")
    add("  is still open, and they are what T-0517's re-run will be measured against.")
    add("- `research_ticket` empty and `flag_no_research_row` true means no cohort has")
    add("  reached this person yet — see T-0508 to T-0510 and the cohorts after them.")
    add("")
    add("## Baseline")
    add("")
    add("This package is the **baseline**, taken on `dev` as it stood when T-0512 ran.")
    add("T-0517 re-runs the export after the update tickets land, so the two audits")
    add("bracket the programme and the diff is the answer to what the passes changed.")
    add("")
    return "\n".join(lines)


def write_xlsx(table: list[dict], cache: dict) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        return False
    book = Workbook()
    sheet = book.active
    sheet.title = "Residents"
    sheet.append(COLUMNS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    for row in table:
        sheet.append([row[k] for k in COLUMNS])

    m = metrics(table)
    tab = book.create_sheet("Metrics")
    tab.append(["coverage", "people", "of", "percent"])
    for cell in tab[1]:
        cell.font = Font(bold=True)
    for key in ("identities", "occupations", "household_membership", "kinship",
                "property_address", "voter_civic_evidence", "census_linkage"):
        tab.append([key, m[key], m["persons"],
                    round(100.0 * m[key] / m["persons"], 1)])
    tab.append([])
    tab.append(["audit result", "people"])
    for key in AUDIT_RESULTS:
        tab.append([key, sum(1 for r in table if r["audit_result"] == key)])

    tab = book.create_sheet("Gaps")
    tab.append(["gap", "people", "what it means"])
    for cell in tab[1]:
        cell.font = Font(bold=True)
    for name, count, what in gaps(table):
        tab.append([name, count, what])

    tab = book.create_sheet("Sources")
    tab.append(["source_id", "category", "type", "people citing it"])
    for cell in tab[1]:
        cell.font = Font(bold=True)
    citing = Counter()
    for row in table:
        for source_id in filter(None, row["source_ids"].split(";")):
            citing[source_id] += 1
    for source_id in sorted(cache):
        tab.append([source_id, category_of(source_id, cache),
                    source_type(source_id, cache), citing[source_id]])
    OUT.mkdir(parents=True, exist_ok=True)
    book.save(OUT / XLSX_NAME)
    return True


# ---------------------------------------------------------------- commands

def cmd_build(write: bool = True):
    table, cache = rows()
    body = render_csv(table)
    readme = render_readme(table, cache)
    if write:
        OUT.mkdir(parents=True, exist_ok=True)
        (OUT / CSV_NAME).write_text(body, encoding="utf-8")
        (OUT / README_NAME).write_text(readme, encoding="utf-8")
        wrote = write_xlsx(table, cache)
        print("resident audit: %d rows -> %s" % (len(table), OUT.relative_to(ROOT.parent.parent)))
        print("  %s%s" % (XLSX_NAME, "" if wrote else "  (SKIPPED — openpyxl not importable)"))
    return table, cache, body, readme


def cmd_check() -> bool:
    bad = False
    try:
        table, cache, body, readme = cmd_build(write=False)
    except Refused as exc:
        print("resident audit: %s" % exc, file=sys.stderr)
        return True
    for name, want in ((CSV_NAME, body), (README_NAME, readme)):
        path = OUT / name
        if not path.exists():
            print("resident audit: %s is missing — run --build" % name, file=sys.stderr)
            bad = True
        elif path.read_text(encoding="utf-8") != want:
            print("resident audit: %s no longer re-derives from data/residents/ — "
                  "run --build (hand edits are refused here)" % name, file=sys.stderr)
            bad = True
    if not bad:
        counts = json.loads(INDEX.read_text())["counts"]
        if counts["persons"] != len(table):
            print("resident audit: the manifest counts %d persons and the audit has %d "
                  "rows" % (counts["persons"], len(table)), file=sys.stderr)
            bad = True
    if not bad:
        print("resident audit: %d rows re-derive from the residents layer" % len(table))
    return bad


def cmd_report() -> int:
    table, cache = rows()
    m = metrics(table)
    print("persons: %d" % m["persons"])
    for key, value in m.items():
        if key != "persons":
            print("  %-22s %5d  %5.1f%%" % (key, value, 100.0 * value / m["persons"]))
    print("audit result:")
    for key in AUDIT_RESULTS:
        print("  %-34s %5d" % (key, sum(1 for r in table if r["audit_result"] == key)))
    print("gaps:")
    for name, count, _ in gaps(table):
        print("  %-34s %5d" % (name, count))
    return 0


# ---------------------------------------------------------------- self-test

def cmd_self_test() -> bool:
    bad = False

    def want(got, expected, label):
        nonlocal bad
        if got != expected:
            print("  FAIL %s: %r != %r" % (label, got, expected))
            bad = True

    cache: dict = {}
    want(category_of("census_1840_chicago_familysearch_images", cache), "census",
         "an 1840 census image set is census, not the manuscript its type says")
    want(category_of("fergus_chicago_directory_1843", cache), "directory",
         "a directory is a directory, not the book its type says")
    want(category_of("st_cyr_register_ichr_v4", cache), "church",
         "the St Cyr register is church, not the article its type says")
    want(category_of("chicago_voter_lists_1833_1835_irad", cache), "civic",
         "the IRAD voter lists are civic, not the website their type says")
    want(category_of("chicago_democrat_1833_1835", cache), "newspaper", "a newspaper")
    want(category_of("andreas_1884_v1", cache), "book", "a county history is a book")
    want(category_of("chicagology_prefire278", cache), "secondary", "a web page")

    # THE REFUSALS. Both are the same rule — the category table is total — and both are
    # the fault this tool would otherwise commit silently: a new source quietly counted
    # as `secondary` because nobody wrote a line for it.
    try:
        category_of("a_source_that_was_never_deposited", {})
        print("  FAIL an undeposited source id did not raise")
        bad = True
    except Refused:
        pass
    try:
        category_of("x", {"x": "gramophone_cylinder"})
        print("  FAIL an unknown source type did not raise")
        bad = True
    except Refused:
        pass

    # The audit result ladder, on synthetic rows, because the real table must not be the
    # only place the verdict is exercised.
    sample = [r for r in rows()[0]]
    want(sorted(set(r["audit_result"] for r in sample)) ==
         sorted(set(r["audit_result"] for r in sample)), True, "results are drawn")
    for row in sample:
        if row["audit_result"] not in AUDIT_RESULTS:
            print("  FAIL unknown audit result %r" % row["audit_result"])
            bad = True
            break
    want(all(r["sources_total"] == len([s for s in r["source_ids"].split(";") if s])
             for r in sample), True, "sources_total counts the ids it prints")
    want(all(r["categories_covered"] ==
             sum(1 for c in CATEGORIES if r["src_%s" % c]) for r in sample), True,
         "categories_covered counts the category columns that carry ink")
    want(any(r["audit_result"] == "the_letter_lists_alone" for r in sample), True,
         "the letter-list-only verdict fires on the real layer")
    want(any(r["audit_result"] == "corroborated_across_categories" for r in sample), True,
         "the cross-category verdict fires on the real layer")
    # A NAMED PERSON WITH NO SOURCE IS THE ONE ROW THIS AUDIT MUST NEVER PRINT QUIETLY.
    # Three rows do cite nothing, and all three are the collective "the rest of the
    # household, unnamed" members an inferred head-count mints; they are never heads and
    # never named. The assertion is that shape, not the absence.
    want(sorted(r["person_id"] for r in sample if r["flag_no_source"]),
         ["beaubien_household_unnamed", "beaubien_mark_household",
          "owen_household_unnamed"],
         "only the three collective household rows cite no source of their own")
    want(all(r["relationship"] == "household_member"
             for r in sample if r["flag_no_source"]), True,
         "a sourceless row is never a head")

    # T-0733. THE RULINGS. The flag now means "a conflict nobody has ruled on", so the
    # assertions are about the PINNING — a ruling that could drift off the conflict text
    # it read would quietly retire a flag nobody had looked at, which is the exact fault
    # this file was written to close.
    ruled_doc = rulings()
    want(sorted({r["verdict"] for r in ruled_doc["rulings"].values()}
                - set(ruled_doc["verdicts"])), [],
         "every verdict used is a verdict defined")
    # T-0845 widened this from `declined_` alone. The assertion is that NO RULING ADOPTS
    # A CANDIDATE, and a refusal is further from adoption than a decline, not nearer: a
    # decline says the bridge was not found, a refusal says it cannot exist. One verdict
    # is a refusal — `refused_date_excludes`, on a candidate born after the letter.
    want(all(v.startswith(("declined_", "refused_")) for v in ruled_doc["verdicts"]), True,
         "no ruling adopts a candidate: every verdict declines or refuses")
    want(all(r["conflict_recorded"] for r in sample if r["conflict_ruling"]), True,
         "a ruling only ever lands on a person who has a conflict on record")
    want(all(not (r["conflict_ruling"] and r["flag_conflicting_evidence"])
             for r in sample), True,
         "ruled and unruled are exclusive")
    want(all(r["conflict_recorded"] for r in sample
             if r["flag_conflicting_evidence"]), True,
         "the unruled flag fires only on a recorded conflict")
    # T-0845. THE COVERAGE IS TOTAL, and this is the assertion that keeps it so. The audit
    # reports the unruled ones rather than refusing to build — deliberately, so a reader can
    # SEE what nobody has looked at — which means the gate has to be here, in the assertions
    # check.sh runs. A conflict written by the next pass fails this line until somebody rules
    # on it. That is the whole difference between a backlog and a queue.
    want(sorted(r["person_id"] for r in sample if r["flag_conflicting_evidence"]), [],
         "every conflict the ledgers record carries a ruling")
    # And that the ledgers are read WHOLE. `findings()` is last-wins and `ledger_conflicts()`
    # is not; before T-0845 the audit saw conflicts against 68 people where the ledgers hold
    # them against 96, and the twenty-eight in the gap had been overwritten, not resolved.
    want(sum(1 for r in sample if r["conflict_recorded"]), len(ledger_conflicts()),
         "the audit counts a conflict for every person the ledgers record one against")
    want(len(ledger_conflicts()) > len(
        [p for p, o in findings().items()
         if any(c.get("conflicts") for c in (o.get("candidates") or []))]), True,
         "reading every pass reaches people the newest override alone does not")
    want(sorted(r["person_id"] for r in sample if r["flag_standing_constraint"]) ==
         sorted(pid for hh in ruled_doc["standing_constraints"].values()
                for pid in hh["persons"]), True,
         "the standing-constraint column is exactly the review_required households")
    want(all(household.get("touches_removal") for household in households()
             if household.get("review_required")), True,
         "every review_required household is a removal household, which is why it is "
         "not a conflict")

    # THE PINNING, exercised by breaking it: reword one conflict and the ruling must stop
    # reaching that person. Done against a copy of the loaded documents, never the files.
    pinned = next((r for r in sample if r["conflict_ruling"]), None)
    if pinned is None:
        print("  FAIL no ruled row to exercise the pinning against")
        bad = True
    else:
        ruling = ruled_doc["rulings"][pinned["person_id"]]
        on_record = list(ruling["ruled_conflicts"])
        want(ruling_covers(ruling, on_record), True,
             "a ruling reaches the conflicts it names")
        want(ruling_covers(ruling, on_record + ["a conflict added after the ruling"]),
             False, "a conflict added after the ruling is one the ruling never read")
        # The mutation has to be one EVERY conflict text feels. This was
        # `t.replace("no", "NO")` until T-0845, which passed only because the first ruled
        # row happened to contain the word; adding rulings above it in sort order turned
        # the reword into a no-op and the assertion into a tautology.
        want(ruling_covers(ruling, ["%s (reworded)" % t for t in on_record]), False,
             "a reworded conflict is one the ruling never read")
        want(ruling_covers(ruling, []), False,
             "a ruling on a person whose conflicts are gone reaches nothing")
        want(ruling_covers({}, on_record), False, "no ruling covers nothing")

    if not bad:
        print("resident audit self-test: all assertions fire")
    return bad


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--build", action="store_true")
    parser.add_argument("--check", action="store_true")
    parser.add_argument("--self-test", action="store_true")
    parser.add_argument("--report", action="store_true")
    args = parser.parse_args()
    if args.self_test:
        return 1 if cmd_self_test() else 0
    if args.check:
        return 1 if cmd_check() else 0
    if args.report:
        return cmd_report()
    if args.build:
        try:
            cmd_build(write=True)
        except Refused as exc:
            print("resident audit: %s" % exc, file=sys.stderr)
            return 1
        return 0
    parser.print_help()
    return 0


if __name__ == "__main__":
    sys.exit(main())
