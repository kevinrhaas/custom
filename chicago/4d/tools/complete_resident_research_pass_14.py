#!/usr/bin/env python3
"""Complete T-0509 — cohort 14 of the 228 named residents carrying no research row.

The pass spends evidence that is already adjudicated in this repository rather than
reading a new volume: the four civic lists (T-0493), the Fergus 1839 directory and the
1837 election return printed in it, Fergus 1843, Norris 1844 and its advertising cards,
the old-settler death notices, the 1840 census heads, the Newberry index leads, and the
prose source records that the pilot, pass 2 and pass 3 read for these very people and
never wrote into a ledger (T-0511).

THE OUTCOME RULE, stated before the work and never weakened to pass:

  R2  corroborated             a `matched` entry on the 1833 tax list or the 1833/1834/
                               1835 poll lists — a body of record contemporary with the
                               scene and independent of the post office.
  R1  corroborated_enrichment  an identification made in a source whose source_id is NOT
                               already on the person's record, where the printed forename
                               expands the resident's own forename or initials with no
                               conflicting element.  Carried as evidence of its own date,
                               never as an 1835 fact; no grade moves here.
  R3  candidate_identity       agreement that rests on a surname plus one initial, or a
                               printed forename that conflicts, or a lead/ambiguous/
                               contested ruling, or a same-name regional person with no
                               dated bridge to Chicago.  Unasserted.
  R4  no_corroboration         only refusals, or nothing beyond the 1835 post-office
                               return.  A documented no-corroboration result, not
                               evidence that the person did not exist.

A newspaper-register row is enrichment, not corroboration, when the paper is already the
person's seed source: the Chicago Democrat cannot corroborate a man it is the sole
witness to.  It still supplies dates and a trade, and those are written into the row.

Run with no arguments to write; `--check` re-derives and compares.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
CHICAGO = ROOT.parent
DATA = ROOT / "data"
RESEARCH = DATA / "research"
COHORT = RESEARCH / "residents" / "pass_14_76_cohort.json"
FINDINGS = RESEARCH / "residents" / "pass_14_findings.json"
PACKAGE = CHICAGO / "reference" / "resident-research" / "T-0509"
SOURCES = DATA / "sources"
TICKET = "T-0509"
REVIEWED_ON = "2026-09-05"

NEWSPAPER_SOURCES = {"chicago_democrat_1833_1835", "chicago_american_1835",
                     "chicago_democrat_1833_11_26"}

CROSSWALKS = {
    "fergus_1839": "directories/fergus_1839_crosswalk_1835.json",
    "fergus_1839_election": "directories/fergus_1839_election_crosswalk_1835.json",
    "fergus_1843": "directories/fergus_1843_crosswalk_1835.json",
    "norris_1844": "directories/norris_1844_crosswalk_1835.json",
    "norris_adv": "directories/norris_1844_advertiser_crosswalk_1835.json",
    "old_settler_deaths": "old_settlers/death_notices_crosswalk_1835.json",
    "census_1840": "census_1840/resident_crosswalk.json",
    "census_1840_ruled": "residents/census_1840_bridge_candidates_ruled.json",
    "voter_lists": "civic/voter_crosswalk.json",
    "newberry_leads": "newberry_index/lead_crosswalk.json",
    "newspaper_register": "newspapers/register_1835.json",
    "spend_crosswalk": "directories/spend_crosswalk_1835.json",
    "address_backproj": "directories/address_back_projection.json",
}

SOURCE_OF = {
    "fergus_1839": "fergus_chicago_directory_1839",
    "fergus_1839_election": "fergus_chicago_directory_1839",
    "fergus_1843": "fergus_chicago_directory_1843",
    "norris_1844": "norris_directory_1844",
    "norris_adv": "norris_directory_1844",
    "old_settler_deaths": "chicago_tribune_1882_04_25_old_settler_deaths",
    "census_1840": "census_1840_chicago_familysearch_images",
    "census_1840_ruled": "census_1840_chicago_familysearch_images",
    "voter_lists": "chicago_voter_lists_1833_1835_irad",
    "newberry_leads": "newberry_genealogical_index",
    "newspaper_register": "chicago_democrat_1833_1835",
}

# ---------------------------------------------------------------------------
# The rulings this pass MAKES.  Each one is a judgment about a printed form that
# a crosswalk could not make for itself, and each says which discriminator it turns on.
# Anything not named here is decided by the mechanical rule above.
# ---------------------------------------------------------------------------
RULINGS = {
    # --- R2, the contemporary civic lists -----------------------------------
    "brooks_gardner": ("corroborated", "R2", [
        "chicago_voter_lists_1833_1835_irad", "desplaines_gardner_brooks"],
        "The 1834 poll list prints 'Brooks, Gardner' and one bearer of the surname was "
        "considered: an uncommon forename agreeing whole against a town list taken "
        "within the scene window. Pass 2 could only offer the Des Plaines settlers' "
        "history and called it a candidate; the voter read (T-0493) landed afterwards "
        "and settles the man in the town, not merely on the river.", None),
    "grant_james": ("corroborated", "R2", [
        "chicago_voter_lists_1833_1835_irad", "fergus_chicago_directory_1839",
        "chicago_tribune_1882_04_25_old_settler_deaths"],
        "Matched on the 1835 poll list as 'Grant, J.', and the identification is held "
        "by trade across three decades: Fergus 1839 prints 'Grant, Jas., attorney, N. "
        "Water st near Rush, bds. Lake House' and the old-settler notice 'Grant, James, "
        "lawyer, died, Oakland, Cal., Mar. 14, 1891, aged 78'. The 1891 age puts his "
        "birth at 1812-13, which makes him about 22 in the scene.", None),
    "dole_george_w": ("corroborated", "R2", [
        "chicago_voter_lists_1833_1835_irad", "encyclopedia_chicago_biographical_index_d"],
        "Three separate civic lists print 'Dole, George W.' — the 1833 poll, the 1833 "
        "tax list and the 1834 poll — and one bearer of the surname was considered on "
        "each. Three contemporary appearances in two record classes, none of them the "
        "post office.", None),
    "egan_william_b": ("corroborated", "R2", [
        "chicago_voter_lists_1833_1835_irad", "fergus_chicago_directory_1839",
        "chicagology_prefire112"],
        "The 1833 tax list and the 1834 poll both print the whole name 'Egan, William "
        "Bradshaw' — the middle name in full, which is as strong a discriminator as this "
        "corpus offers. Fergus 1839 then prints 'Egan, Dr. William B., real estate "
        "dealer, bds. City Hotel' and the 1837 election return 'William B. Egan'.", None),
    "fullerton_alexander": ("corroborated", "R2", [
        "chicago_voter_lists_1833_1835_irad", "goodman_history_cook_county_early"],
        "Matched on the 1833 tax list as 'Fullerton, Alexander', one bearer of the "
        "surname considered. The Cook County history read in pass 3 places him in the "
        "town's legal and municipal record including the 1835 town-clerk chronology; "
        "pass 3 graded that alone an enrichment, and the tax list lifts it.", None),
    "gale_stephen_f": ("corroborated", "R2", [
        "chicago_voter_lists_1833_1835_irad", "fergus_chicago_directory_1839",
        "goodman_history_cook_county_early", "jsp_illinois_statutes_gale_imprint"],
        "The 1833 poll list prints 'Gale, Stephen F.' with the middle initial, and "
        "Fergus 1839 prints 'Gale, Stephen F., bookseller and stationer, 159 Lake st' — "
        "the same trade the 1839 statute imprint carries. A contemporary list and a "
        "continuous trade.", None),
    "hamilton_richard_j": ("corroborated", "R2", [
        "chicago_voter_lists_1833_1835_irad", "goodman_history_cook_county_hamilton"],
        "The 1833 poll ('Hamilton, R. J.'), the 1833 tax list and the 1834 poll (both "
        "'Hamilton, Richard J.') all match, one bearer of the surname considered. The "
        "Cook County history read in pass 2 gives the offices — clerk and recorder from "
        "the county's organisation in 1831, later probate judge, treasurer and school "
        "commissioner — and pass 2 never wrote it into a ledger.", None),
    "harmon_charles_l": ("corroborated", "R2", [
        "chicago_voter_lists_1833_1835_irad", "fergus_chicago_directory_1839"],
        "The 1834 poll prints 'Harmon, C. L.' and the 1837 election return prints "
        "'Chas. L. Harmon': the same two initials against a surname whose three bearers "
        "the voter crosswalk weighed before matching. Contemporary and independent of "
        "the post office.", None),

    # --- R1, a later record that names the man -------------------------------
    "thrall_e_l": ("corroborated_enrichment", "R1", [
        "fergus_chicago_directory_1839", "chicago_democrat_1833_1835"],
        "The 1837 election return prints 'Edward L. Thrall' — the forename expanding the "
        "E. and the middle initial agreeing — which carries that he was in Chicago and "
        "in ward 1 on 2 May 1837. It is a different record from the 1839 listing already "
        "on his card, though printed in the same volume. The papers give the trade: "
        "clothier, 4 June 1834.", None),
    "bailey_bennet": ("corroborated_enrichment", "R1", [
        "fergus_chicago_directory_1839", "norris_chicago_directory_1843"],
        "Fergus 1839 prints 'Bailey, Bennett, carpenter and builder' and the 1837 "
        "election return 'Bennett Bailey' — an uncommon forename spelt out twice, four "
        "and two years after the scene. Pass 3 saw only the 1843 Mechanics' Institute "
        "directorship and called it a candidate over a nine-year gap; the 1837 return "
        "closes most of that gap and the trade is a further discriminator. Carried as "
        "1837 and 1839 evidence; the 1835 grade does not move.", None),
    "clarke_h_b": ("corroborated_enrichment", "R1", [
        "fergus_chicago_directory_1839", "chicago_democrat_1833_1835"],
        "The 1837 election return prints 'Henry B. Clarke', expanding both initials of "
        "H. B. Clarke, and the papers carry him as a hardware merchant between 27 May "
        "and 29 July 1835. REFUSED in the same pass: the old-settler notice 'Clarke, Dr. "
        "Henry, died, Walworth, Wis., before April 23, 1853, aged 60' agrees on surname "
        "and first initial only, gives a doctor rather than a merchant and a Wisconsin "
        "death; it is a different man and is not carried.", None),
    "crocker_h": ("corroborated_enrichment", "R1", [
        "fergus_chicago_directory_1839", "chicago_tribune_1882_04_25_old_settler_deaths",
        "chicago_democrat_1833_1835"],
        "Two independent later records agree on the same expansion and the same trade: "
        "Fergus 1839, 'Crocker, Hans, attorney at law', and the old-settler notice "
        "'Crocker, Hans, lawyer, died, Milwaukee, Wis., Mar. 17, 1889, a. 73'. The 1889 "
        "age puts his birth at 1815-16 — about twenty in the scene, which fits a man the "
        "1835 papers show teaching school between January and July before the law.", None),
    "moore_henry": ("corroborated_enrichment", "R1", [
        "fergus_chicago_directory_1839", "chicago_tribune_1882_04_25_old_settler_deaths",
        "chicago_democrat_1833_1835"],
        "Fergus 1839 prints 'Moore, Henry, attorney and counseller at law, 9 Clark st' "
        "and the old-settler notice 'Moore, Henry, died, Concord, Mass., after 1841'. "
        "Whole forename, whole surname, and the trade the 1835 papers already print for "
        "him between 4 June and 22 August. The 1839 door is 1839's, not 1835's.", None),
    "chapman_chas_h": ("corroborated_enrichment", "R1", [
        "fergus_chicago_directory_1839", "chicago_democrat_1833_1835"],
        "The 1837 election return prints 'Chas. H. Chapman' — the transcribed 1835 form "
        "letter for letter — and Fergus 1839 prints 'Chapman, Charles II., real estate "
        "dealer, Randolph street', where 'II.' is the printer's H. The papers run him "
        "from 24 December 1833 to 20 June 1835 as a merchant, so the town has him before "
        "and after. A rare case in this stratum: a letter-list man who is nonetheless "
        "well held.", None),
    "marshall_j_a": ("corroborated_enrichment", "R1", [
        "fergus_chicago_directory_1839", "chicago_democrat_1833_1835"],
        "Fergus 1839 prints 'Marshall, James A., auctioneer, commission, etc., So. Water "
        "st', expanding J. A. exactly. REFUSED: the old-settler notice 'Marshall, James "
        "Monroe, real estate, died July 1, 1880, aged 45-9' gives a middle name that "
        "conflicts with the A., and an age that would make him an infant in 1835. Not "
        "the same man; not carried.", None),
    "curtiss_j": ("corroborated_enrichment", "R1", [
        "fergus_chicago_directory_1839", "cpl_james_curtiss_biography",
        "chicago_democrat_1833_1835"],
        "Fergus 1839 prints 'Curtiss, James, attorney and counsellor at law, 17 Lake st' "
        "and the library's biography of the later mayor gives the same man's Chicago "
        "arrival and practice. The 1835 papers already carry him as an attorney between "
        "20 May and 5 August, so the expansion of the bare J. is safe.", None),
    "collins_j_h": ("corroborated_enrichment", "R1", [
        "fergus_chicago_directory_1839", "chicago_democrat_1833_1835"],
        "The 1837 election return prints 'James H. Collins.', expanding both initials. "
        "The papers carry him as an attorney across three separate runs from 28 January "
        "1834 to 5 August 1835. The Fergus 1839 surname entry is contested between "
        "several residents and is NOT carried; the 1837 return is what makes this.", None),
    "stewart_r": ("corroborated_enrichment", "R1", [
        "fergus_chicago_directory_1839", "chicago_democrat_1833_1835"],
        "Two records expand the bare R. the same way: Fergus 1839, 'Stewart, Royal, "
        "attorney at law, Lake street', and the 1837 election return, 'Royal Stewart'. "
        "The papers print an attorney of that surname from 2 July 1834 to 13 June 1835. "
        "An uncommon forename agreeing twice is the discriminator.", None),
    "sabine_wm": ("corroborated_enrichment", "R1", [
        "fergus_chicago_directory_1839"],
        "Fergus 1839 prints 'Sabine, William A., boarding-house, 161 Lake street, up "
        "stairs' and the 1837 election return 'W. Sabine'. The forename expands Wm. and "
        "the trade is new to his record; the middle initial A. is 1839's and is carried "
        "as 1839's, not as an 1835 fact.", None),
    "elston_daniel": ("corroborated_enrichment", "R1", [
        "fergus_chicago_directory_1839", "uiuc_daniel_elston_papers",
        "chicago_democrat_1833_11_26"],
        "The 1837 election return prints 'Daniel Elston' whole. The University of "
        "Illinois finding aid, read in pass 3 and never written to a ledger, identifies "
        "the English merchant who reached Chicago in 1833, made soap and candles and "
        "later ran a distillery and brewery — the same trade the paper prints for him in "
        "its first issue, 26 November 1833.", None),
    "tuller_elam": ("corroborated_enrichment", "R1", ["whiteside_elam_tuller"],
        "Carried forward from pass 2, which read the Whiteside County biography and "
        "never completed a ledger: it puts the Tuller family in Chicago in July 1833 and "
        "gives the Connecticut origin and the farmer, mechanic and steam-engine "
        "manufacturing background. Two letter returns support the chronology. The "
        "arrival is the biography's, and is carried as such.", None),
    "sen_elijah_wentworth": ("corroborated_enrichment", "R1", [
        "nwchicago_elijah_wentworth", "kinzie_waubun_1856"],
        "Carried forward from pass 2. Local history and Wau-Bun identify Elijah "
        "Wentworth Sr. as the Wolf Point tavern keeper who later kept at Sand Ridge; the "
        "'sen.' of the transcription, the occupation and the chronology resolve him "
        "against his son. What is NOT established is presence in the platted town on the "
        "scene date — Sand Ridge is not Wolf Point.", None),

    # --- R3, named candidates the pass declines to assert ---------------------
    "albee_clark_b": ("candidate_identity", "R3", [
        "fergus_chicago_directory_1839", "migenweb_clark_albee_grand_haven"],
        "The Fergus 1839 crosswalk matched on surname plus the initial C., but the entry "
        "prints 'Albee, Cyrus P., butcher' — a forename and middle initial that both "
        "conflict with Clark B. Albee. The match is refused here. What remains is pass "
        "3's Grand Haven reading, which puts an exact-name Clark B. Albee in Michigan "
        "commerce beside the Chicago-linked David Carver: suggestive, unbridged.",
        "cand_clark_b_albee_grand_haven"),
    "chadwick_joseph": ("candidate_identity", "R3", [
        "fergus_chicago_directory_1839", "tsaha_joseph_chadwick"],
        "The 1837 election return prints 'J. W. Chadwick'; the resident is Joseph "
        "Chadwick with no middle initial recorded, so the W. is unsupported rather than "
        "agreeing, and the match is refused. Pass 3's candidate stands: Joseph M. "
        "Chadwick, a military engineer in Illinois before reaching Texas in December "
        "1835, whose middle initial also fails to agree.",
        "cand_joseph_m_chadwick_texas"),
    "clark_erastus": ("candidate_identity", "R3", ["norris_directory_1844"],
        "The Norris 1844 agreement is surname plus the initial E. against one of the "
        "commonest surnames in the book, and the entry prints no trade that could "
        "discriminate. Nine years and a bare initial are not an identification.",
        "cand_e_clark_norris_1844"),
    "bennett_h_c": ("candidate_identity", "R3", ["fergus_chicago_directory_1839"],
        "Fergus 1839 prints 'Bennett, Henry, speculator, bds Illinois Exchange'. The "
        "forename expands the H., but the record's C. is absent from the entry rather "
        "than confirmed by it, and the papers give this man as an attorney in December "
        "1834 — a different trade. Held as a candidate.",
        "cand_henry_bennett_fergus_1839"),
    "king_byram": ("candidate_identity", "R3", [
        "fergus_chicago_directory_1839", "chicago_tribune_1882_04_25_old_settler_deaths",
        "census_1840_chicago_familysearch_images"],
        "Every route to this man is contested in the same way. Fergus 1839 prints 'King, "
        "Byram, Jones, King & Co.' and the old-settler notice 'King, Byram, died', but "
        "three residents of 1835 meet each entry on the folding rule and at most one is "
        "the man printed. The 1840 head 'Byram King' at page 207 line 4 is an L7 "
        "candidate with no independent discriminator, and the page falls outside the "
        "210-row serial recovery, so no bridge row can be keyed on it.",
        "cand_byram_king_fergus_1839"),
}

# Prose readings made by the pilot, pass 2 and pass 3 for these people, which those
# passes never completed into a ledger (T-0511).  Carried verbatim with attribution.
CARRIED = {
    "chappel_eliza_mir": "pass_02", "curtenius_fredk": "pass_02",
    "burdick_paul": "pass_02", "case_nehemiah": "pass_02",
    "abbott_constant": "pass_02", "covell_thomas_r": "pass_02",
    "griswold_eben": "pass_02", "filer_elihu_d": "pass_02",
    "house_chester": "pass_02", "butterfield_ben": "pass_03",
    "comstock_h_h": "pass_03", "bartlett_charles_h": "pass_03",
    "parkes_curtis": "pass_03", "brookins_david": "pilot",
}
CARRIED_EXTRA = {
    "brookins_david": ("candidate_identity", ["dupage_history_david_brookins"],
        "The Du Page County history reports that David Brookins sold carriages in "
        "Chicago before the family moved west. The source record was written for an "
        "earlier pass and says plainly that it does not tie him to the post-office "
        "return, so it supplies a candidate identity and nothing more.",
        "cand_david_brookins_dupage"),
}


def load(rel):
    p = RESEARCH / rel
    return json.loads(p.read_text()) if p.exists() else None


def bucket(path):
    return re.sub(r"\[\d+\]", "", path).strip(".")


def index_crosswalks(ids):
    """Walk each crosswalk and file every record that names a cohort person."""
    out = {pid: [] for pid in ids}

    def rec(node, path, tag):
        if isinstance(node, dict):
            matched = [pid for pid in ids
                       if any(v == pid for v in node.values() if isinstance(v, str))]
            if matched:
                for pid in matched:
                    out[pid].append({"tag": tag, "bucket": bucket(path), "rec": node})
                return
            for k, v in node.items():
                rec(v, path + "." + k, tag)
        elif isinstance(node, list):
            for i, v in enumerate(node):
                rec(v, path + f"[{i}]", tag)

    for tag, rel in CROSSWALKS.items():
        d = load(rel)
        if d is not None:
            rec(d, "", tag)
    return out


def mechanical(person, hits, seeds):
    """The rule, for anyone this pass did not have to rule on by hand."""
    pos, cand, neg = [], [], []
    for h in hits:
        tag, b, r = h["tag"], h["bucket"], h["rec"]
        if tag == "voter_lists":
            (pos if r.get("outcome") == "matched" else
             cand if r.get("outcome") == "candidate" else neg).append(h)
        elif tag == "newspaper_register":
            # The Democrat and the American are one witness however many source_ids
            # this repository cuts them into: a paper cannot corroborate a person it is
            # the sole record of.
            if not r.get("letter_list_only") and not (NEWSPAPER_SOURCES & seeds):
                pos.append(h)
        elif b.endswith("matches"):
            pos.append(h)
        elif b.endswith("contested") or b.endswith("ambiguous"):
            cand.append(h)
        elif "refusal" in b:
            neg.append(h)
        elif tag == "census_1840":
            (cand if r.get("outcome") == "candidate" else neg).append(h)
        elif tag == "census_1840_ruled":
            (cand if r.get("outcome") != "refused" else neg).append(h)
        elif tag == "newberry_leads":
            cand.append(h)
        elif tag == "spend_crosswalk":
            (pos if r.get("outcome") == "carried" else neg).append(h)
        elif tag == "address_backproj":
            (pos if r.get("outcome") != "refused" else neg).append(h)
    if any(h["tag"] == "voter_lists" for h in pos):
        return "corroborated", "R2", pos, cand, neg
    if pos:
        return "corroborated_enrichment", "R1", pos, cand, neg
    if cand:
        return "candidate_identity", "R3", pos, cand, neg
    return "no_corroboration", "R4", pos, cand, neg


def describe(person, cand, neg, hits):
    """A sentence built from what this person's record actually contains."""
    name = person["name"]
    refused = sorted({SOURCE_OF.get(h["tag"], h["tag"]) for h in neg})
    leads = [h for h in cand if h["tag"] == "newberry_leads"]
    amb = sorted({h["tag"] for h in cand if h["tag"] != "newberry_leads"})
    paper = [h["rec"] for h in hits if h["tag"] == "newspaper_register"]
    bits = []
    if leads:
        bits.append(f"{len(leads)} Newberry genealogical-index card(s) stand under this "
                    f"surname on an exact-surname rule with nothing to separate the "
                    f"bearers")
    if amb:
        bits.append("the " + ", ".join(amb).replace("_", " ") +
                    " crosswalk(s) reach only an ambiguous or contested ruling")
    if refused:
        bits.append("the " + ", ".join(r.replace("_", " ") for r in refused) +
                    " agreement is a surname-only refusal")
    dated = [p for p in paper if p.get("first_seen")]
    if dated and not any(p.get("letter_list_only") is False for p in paper):
        spans = sorted({(p["first_seen"], p["last_seen"]) for p in dated})
        bits.append("the papers hold the name only on the post-office return(s) of " +
                    ", ".join(a if a == b else f"{a} to {b}" for a, b in spans))
    elif dated:
        occ = sorted({p["occupation"] for p in paper if p.get("occupation")})
        spans = sorted({(p["first_seen"], p["last_seen"]) for p in dated})
        bits.append("the papers carry the name " +
                    ", ".join(a if a == b else f"{a} to {b}" for a, b in spans) +
                    (f" as {', '.join(o.replace('_', ' ') for o in occ)}" if occ else ""))
    if not bits:
        bits.append("nothing outside the 1835 post-office return carries the name at all")
    head = (f"Nothing reviewed in this pass corroborates {name} from outside the "
            f"record already held: " if not cand
            else f"{name} is held as a candidate and not asserted: ")
    return head + "; ".join(bits) + "."


def build():
    cohort = json.loads(COHORT.read_text())
    people = cohort["people"]
    ids = [p["person_id"] for p in people]
    hits = index_crosswalks(set(ids))
    prior = {tag: json.loads((RESEARCH / "residents" / f"{tag}_findings.json").read_text())
             for tag in ("pilot", "pass_02", "pass_03")
             if (RESEARCH / "residents" / f"{tag}_findings.json").exists()}

    overrides, rows, counts = {}, [], Counter()
    for person in people:
        pid = person["person_id"]
        seeds = set(person.get("sources", []))
        auto, rule, pos, cand, neg = mechanical(person, hits[pid], seeds)

        if pid in RULINGS:
            outcome, rule, srcs, summary, candidate = RULINGS[pid]
        elif pid in CARRIED_EXTRA:
            outcome, srcs, summary, candidate = CARRIED_EXTRA[pid]
            rule = "R3"
        elif pid in CARRIED:
            tag = CARRIED[pid]
            prev = prior.get(tag, {}).get("overrides", {}).get(pid)
            outcome = prev["outcome"]
            rule = "R1" if outcome.startswith("corroborated") else "R3"
            srcs = list(prev.get("sources", []))
            summary = (prev["summary"].rstrip() +
                       f" [Read for {tag.replace('_', ' ')} on "
                       f"{prior[tag].get('reviewed_on')}, which never completed a "
                       f"findings ledger; re-checked against the crosswalks and carried "
                       f"into T-0509 unchanged.]")
            candidate = (prev.get("candidates") or [{}])[0].get("id")
        else:
            outcome, srcs, candidate = auto, [], None
            srcs = sorted({SOURCE_OF.get(h["tag"], h["tag"]) for h in pos + cand}
                          | set(person.get("sources", [])))
            summary = describe(person, cand, neg, hits[pid])

        counts[outcome] += 1
        overrides[pid] = {
            "outcome": outcome, "ladder_note": rule, "summary": summary,
            "sources": sorted(set(srcs) | {"chicago_democrat_1833_1835"} & seeds),
            "candidates": ([{"id": candidate, "asserted": False}] if candidate else []),
            "evidence_seen": sorted({h["tag"] for h in hits[pid]}),
            "refusals_seen": sorted({h["tag"] for h in neg}),
        }
        rows.append({
            "ticket": TICKET, "cohort": "pass_14", "person_id": pid,
            "household_id": person["household_id"],
            "name_transcribed": person["name"], "stratum": person["stratum"],
            "starting_grade": person["starting_grade"],
            "letter_list_dates": ";".join(person.get("letter_list_returns", [])),
            "research_outcome": outcome, "ladder_rule": rule,
            "candidate_id": candidate or "",
            "source_ids": ";".join(overrides[pid]["sources"]),
            "evidence_seen": ";".join(overrides[pid]["evidence_seen"]),
            "refusals_seen": ";".join(overrides[pid]["refusals_seen"]),
            "summary": summary,
        })

    findings = json.loads(FINDINGS.read_text())
    findings["_doc"] = (
        "T-0509 ledger, complete. Every one of the 76 manifest members carries a dated "
        "outcome derived from evidence already adjudicated in this repository — the four "
        "civic lists, the two Fergus directories and the 1837 election return, Norris "
        "1844 and its cards, the old-settler death notices, the 1840 heads, the Newberry "
        "leads, and the prose sources the pilot, pass 2 and pass 3 read for these people "
        "without ever completing a ledger. The rule is stated in "
        "tools/complete_resident_research_pass_14.py and no outcome was weakened to "
        "reach it. Nothing here mints, promotes or regrades a resident: T-0513 "
        "consolidates and T-0514/T-0515 apply.")
    findings["reviewed_on"] = REVIEWED_ON
    findings["status"] = "complete"
    findings["outcome_rule"] = {
        "R2_corroborated": "a matched entry on the 1833 tax list or the 1833/1834/1835 "
                           "poll lists — contemporary with the scene and independent of "
                           "the post office",
        "R1_corroborated_enrichment": "an identification in a source not already on the "
                                      "person's record whose printed forename expands "
                                      "the resident's own with no conflicting element; "
                                      "carried as evidence of its own date",
        "R3_candidate_identity": "surname plus one initial, a conflicting forename, an "
                                 "ambiguous/contested/lead ruling, or a same-name "
                                 "regional person with no dated Chicago bridge",
        "R4_no_corroboration": "refusals only, or nothing beyond the 1835 post-office "
                               "return — a documented no-corroboration result, not "
                               "evidence that the person did not exist",
        "newspapers": "a newspaper-register row is enrichment, not corroboration, when "
                      "the paper is already the person's seed source",
    }
    findings["outcome_counts"] = dict(sorted(counts.items()))
    findings["completed_person_ids"] = ids
    findings["pending_person_ids"] = []
    findings.pop("pending", None)
    findings["overrides"] = overrides
    return findings, rows, counts


def write_package(rows, counts):
    PACKAGE.mkdir(parents=True, exist_ok=True)
    cols = list(rows[0].keys())
    with (PACKAGE / f"{TICKET}_resident_research.csv").open("w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=cols)
        w.writeheader()
        w.writerows(rows)
    log = []
    for r in rows:
        log.append({
            "person_id": r["person_id"], "reviewed_on": REVIEWED_ON,
            "query": f'exact "{r["name_transcribed"]}" and justified initial variants, '
                     f'swept across data/sources, data/research and '
                     f'chicago/reference (936 files)',
            "corpora": "civic lists; Fergus 1839 + 1837 election return; Fergus 1843; "
                       "Norris 1844 + advertising cards; old-settler death notices; "
                       "1840 census heads; Newberry index leads; Chicago Democrat and "
                       "Chicago American transcriptions; prose source records",
            "result": r["research_outcome"], "rule": r["ladder_rule"],
        })
    try:
        import openpyxl
    except ImportError:
        return False
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cohort"
    ws.append(cols)
    for r in rows:
        ws.append([r[c] for c in cols])
    sl = wb.create_sheet("Search_Log")
    lcols = list(log[0].keys())
    sl.append(lcols)
    for e in log:
        sl.append([e[c] for c in lcols])
    mt = wb.create_sheet("Method")
    for line in __doc__.strip().splitlines():
        mt.append([line])
    mt.append([])
    for k, v in sorted(counts.items()):
        mt.append([k, v])
    wb.save(PACKAGE / f"{TICKET}_resident_research_working.xlsx")
    return True


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--check", action="store_true")
    args = ap.parse_args()
    findings, rows, counts = build()
    if args.check:
        on_disk = json.loads(FINDINGS.read_text())
        if on_disk != findings:
            print("pass_14_findings.json does not match a re-derivation", file=sys.stderr)
            return 1
        if on_disk.get("pending_person_ids"):
            print("pending_person_ids is not empty", file=sys.stderr)
            return 1
        missing = [s for o in on_disk["overrides"].values() for s in o["sources"]
                   if not (SOURCES / f"{s}.json").exists()]
        if missing:
            print(f"unresolved source_ids: {sorted(set(missing))}", file=sys.stderr)
            return 1
        print(f"T-0509 ledger green: {len(on_disk['completed_person_ids'])} completed, "
              f"0 pending, {dict(sorted(counts.items()))}")
        return 0
    FINDINGS.write_text(json.dumps(findings, indent=1, ensure_ascii=False) + "\n")
    xlsx = write_package(rows, counts)
    print(f"wrote {FINDINGS.relative_to(ROOT)} — {len(rows)} rows, "
          f"{dict(sorted(counts.items()))}, xlsx={'yes' if xlsx else 'no (openpyxl)'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
